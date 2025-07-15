#!/usr/bin/env python3
"""
主窗口模块 - OCR识别和标注功能
"""

import os
import re
import sys
import time
import json
import logging
import math
import random
import tempfile
import numpy as np
from datetime import datetime
from typing import List, Dict, Tuple, Set, Optional, Any, Union
from pathlib import Path

from PySide6.QtWidgets import (
    QMainWindow, QGraphicsScene, QGraphicsPixmapItem, 
    QFileDialog, QMessageBox, QLabel, QSlider, QWidget,
    QVBoxLayout, QHBoxLayout, QGridLayout, QSplitter, QToolBar, 
    QLineEdit, QPushButton, QCheckBox, QComboBox, QStatusBar,
    QScrollArea, QGroupBox, QFrame, QFormLayout, QProgressBar,
    QDialog, QApplication, QSizePolicy, QListWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QMenu,
    QSpinBox, QListWidgetItem, QInputDialog
)
from PySide6.QtCore import (
    Qt, QObject, QRunnable, Signal, QThreadPool, QTimer, 
    QSize, QPoint, QPointF, QRectF, QRect, QEvent, QFile,
    Slot, QSettings
)
from PySide6.QtGui import (
    QPixmap, QImage, QColor, QPainter, QPen, QBrush, 
    QFont, QFontMetrics, QKeySequence, QPainterPath, 
    QTransform, QPalette, QIcon, QGuiApplication, QAction,
    QIntValidator
)

# 检查PaddleOCR是否可用
try:
    import paddle
    HAS_OCR_SUPPORT = True
except ImportError:
    HAS_OCR_SUPPORT = False

# 导入自定义模块
from core.annotation_item import BubbleAnnotationItem
from core.file_loader import FileLoader
from ui.graphics_view import GraphicsView
from ui.annotation_list import AnnotationTable
from ui.property_editor import PropertyEditor

from utils.constants import (
    APP_TITLE, FILE_DIALOG_FILTER, DEFAULT_WINDOW_SIZE, DEFAULT_WINDOW_POSITION,
    DEFAULT_OCR_LANGUAGES, PDF_QUALITY_OPTIONS, OCR_TEXT_TYPE_COLORS,
    OCR_TYPE_TO_STYLE, STYLE_NAME_MAP, STYLE_NAME_REVERSE_MAP,
    OCR_FILTER_OPTIONS, OCR_FILTER_TYPE_MAP, UI_COLORS, SUPPORTED_IMAGE_FORMATS,
    SUPPORTED_PDF_FORMATS, SUPPORTED_DXF_FORMATS,
    BUBBLE_SIZE_MIN_PERCENT, BUBBLE_SIZE_MAX_PERCENT, BUBBLE_SIZE_DEFAULT_PERCENT, BUBBLE_SIZE_STEP,
    BUBBLE_REORDER_GRID_SIZE
)
from utils.dependencies import HAS_OCR_SUPPORT, HAS_GPU_SUPPORT, HAS_PADDLE_OCR, get_requirements_message

# 只导入PaddleOCR工作器
if HAS_PADDLE_OCR:
    from core.paddle_ocr_worker import PaddleOCRWorker

# 检查Excel支持
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from copy import copy
    import pandas as pd
    HAS_EXCEL_SUPPORT = True
except ImportError:
    HAS_EXCEL_SUPPORT = False

# 配置日志记录
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('debug.log', 'w', 'utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger('PyQtBubble')

# 添加一个自定义事件类用于从线程传递加载结果到主线程
class LoadPDFEvent(QEvent):
    """自定义事件用于将PDF加载结果从线程传递到主线程"""
    EVENT_TYPE = QEvent.Type(QEvent.registerEventType())
    
    def __init__(self, pixmap, temp_path):
        super().__init__(self.EVENT_TYPE)
        self.pixmap = pixmap
        self.temp_path = temp_path

class PDFLoaderSignals(QObject):
    """PDF加载工作线程的信号类"""
    finished = Signal(QPixmap, str)  # 成功加载后发出信号：pixmap, 临时文件路径
    error = Signal(str)  # 加载出错时发出信号
    progress = Signal(int)  # 加载进度信号

class PDFLoaderWorker(QRunnable):
    """PDF加载工作线程"""
    def __init__(self, pdf_path, page_index, quality=4.0, force_resolution=False):
        super().__init__()
        self.pdf_path = pdf_path
        self.page_index = page_index
        self.quality = quality
        self.force_resolution = force_resolution
        self.signals = PDFLoaderSignals()
        
    def run(self):
        """执行PDF加载"""
        try:
            start_time = time.time()
            logger.debug(f"开始加载PDF页面: {self.page_index+1}, 质量: {self.quality}, 强制分辨率: {self.force_resolution}")
            
            self.signals.progress.emit(10)
            # 创建一个临时场景，用于在线程中处理
            temp_scene = QGraphicsScene()
            logger.debug(f"创建临时场景耗时: {time.time() - start_time:.2f}秒")
            
            self.signals.progress.emit(30)
            
            # 调用FileLoader加载PDF
            logger.debug(f"开始调用FileLoader.load_pdf...")
            load_start = time.time()
            pixmap, temp_path = FileLoader.load_pdf(
                self.pdf_path, temp_scene, self.page_index, quality=self.quality,
                force_resolution=self.force_resolution
            )
            logger.debug(f"FileLoader.load_pdf完成，耗时: {time.time() - load_start:.2f}秒")
            
            self.signals.progress.emit(90)
            
            if pixmap and not pixmap.isNull() and temp_path:
                logger.debug(f"PDF页面加载成功: {self.page_index+1}, 尺寸: {pixmap.width()}x{pixmap.height()}")
                # 成功加载，发送信号
                self.signals.finished.emit(pixmap, temp_path)
            else:
                logger.error(f"PDF页面加载失败: pixmap为空或temp_path为空")
                # 加载失败
                self.signals.error.emit("无法加载PDF页面")
                
            logger.debug(f"PDF加载线程总耗时: {time.time() - start_time:.2f}秒")
                
        except Exception as e:
            logger.exception(f"PDF加载过程中发生异常: {str(e)}")
            # 处理异常
            self.signals.error.emit(str(e))

class FileLoaderSignals(QObject):
    """文件加载工作线程的信号类"""
    finished = Signal(str, QPixmap)  # 成功加载后发出信号：文件路径, 图像数据
    pdf_loaded = Signal(str, int)  # PDF加载成功的信号：文件路径, 页数
    error = Signal(str)  # 加载出错时发出信号
    progress = Signal(int, str)  # 加载进度信号：进度值, 描述

class FileLoaderWorker(QRunnable):
    """文件加载工作线程"""
    def __init__(self, file_path, pdf_quality="高清 (4x)"):
        super().__init__()
        self.file_path = file_path
        self.pdf_quality = pdf_quality
        self.signals = FileLoaderSignals()
        
    def run(self):
        """执行文件加载"""
        try:
            file_path = Path(self.file_path)
            extension = file_path.suffix.lower()
            
            self.signals.progress.emit(10, f"正在加载文件 {file_path.name}...")
            
            # 处理图像文件
            if extension in SUPPORTED_IMAGE_FORMATS:
                self.signals.progress.emit(30, f"正在加载图像 {file_path.name}...")
                pixmap = FileLoader.load_image(str(file_path))
                if pixmap:
                    self.signals.progress.emit(90, "图像加载成功")
                    self.signals.finished.emit(str(file_path), pixmap)
                else:
                    self.signals.error.emit(f"无法加载图像文件: {file_path.name}")
                    
            # 处理PDF文件
            elif extension in SUPPORTED_PDF_FORMATS:
                self.signals.progress.emit(20, f"正在分析PDF文件 {file_path.name}...")
                
                # 获取PDF页数
                page_count = FileLoader.get_pdf_page_count(str(file_path))
                if page_count == 0:
                    self.signals.error.emit("无法读取PDF文件或PDF文件不包含任何页面")
                    return
                
                # 通知PDF加载成功
                self.signals.progress.emit(80, f"PDF文件加载成功，共 {page_count} 页")
                self.signals.pdf_loaded.emit(str(file_path), page_count)
                    
            # 处理DXF文件
            elif extension in SUPPORTED_DXF_FORMATS:
                self.signals.error.emit("DXF文件加载尚未实现多线程支持")
                
            else:
                self.signals.error.emit(f"不支持的文件格式: {extension}")
                
        except Exception as e:
            self.signals.error.emit(f"加载文件时发生错误: {str(e)}")

class MainWindow(QMainWindow):
    """
    主窗口类
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("图纸标注系统")
        self.setWindowIcon(QIcon("assets/icon.ico"))
        
        # 设置初始窗口大小
        self.setMinimumSize(1200, 800)  # 设置最小大小
        self.resize(1600, 900)  # 默认初始大小
        
        # 创建线程池
        self.thread_pool = QThreadPool()
        
        # 初始化对象属性
        self.graphics_scene = QGraphicsScene()
        self.graphics_view = None  # 稍后在setup_ui中创建
        self.property_editor = None  # 稍后在setup_ui中创建
        self.annotation_table = None  # 稍后在setup_ui中创建
        self.force_resolution_checkbox = None # 强制分辨率复选框
        self.current_pixmap = None
        self.current_file_path = ""
        self.annotations = []
        self.current_annotation = None
        self.annotation_counter = 0
        self.is_selecting_area = False
        
        # OCR相关
        self.ocr_worker = None
        self.ocr_results = []
        self.area_ocr_worker = None
        
        # PDF相关
        self.pdf_file_path = ""
        self.pdf_page_count = 0
        self.current_pdf_page = 0
        self.previous_page = 0
        self.pdf_pages_cache = {}  # 页面缓存：{页码: 临时文件路径}
        
        # 多页文档的数据存储
        self.annotations_by_page = {}  # {页码: 标注数据列表}
        self.ocr_results_by_page = {}  # {页码: OCR结果列表}
        
        # 屏蔽区域
        self.masked_regions = []
        self.mask_items = []
        self.is_selecting_mask = False
        
        # 标注颜色
        self.annotation_color = QColor(255, 0, 0, 200)  # 默认红色，半透明
        self.next_annotation_color = None  # 用于存储下一个标注的颜色
        self.next_annotation_scale = 1.0  # 用于存储下一个标注的比例因子
        self.next_annotation_size = None  # 用于存储下一个标注的大小
        
        # 创建半透明加载对话框
        self.create_loading_dialog()
        
        # 设置UI
        self.setup_ui()
        
        # 初始化完成后强制处理事件，确保所有UI元素已正确创建
        QApplication.processEvents()
        
        logger.debug("MainWindow初始化完成")
        
    def create_loading_dialog(self):
        """创建半透明的加载对话框"""
        dialog = QDialog(self, Qt.FramelessWindowHint)
        dialog.setAttribute(Qt.WA_TranslucentBackground)
        dialog.setModal(True)  # 设置为模态对话框，阻止其他交互
        dialog.setStyleSheet("background-color: rgba(0, 0, 0, 180);")
        
        layout = QVBoxLayout(dialog)
        
        # 加载提示标签
        label = QLabel("⏳ 正在加载页面...\n请稍候", dialog)
        label.setStyleSheet("color: white; font-size: 20pt; font-weight: bold; background-color: rgba(0, 0, 0, 200); padding: 40px; border-radius: 20px;")
        label.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(label)
        layout.setAlignment(Qt.AlignCenter)
        
        self.loading_dialog = dialog
        self.loading_label = label
        
        logger.debug("创建了半透明加载对话框")
        
    def setup_ui(self):
        self.setGeometry(*DEFAULT_WINDOW_POSITION, *DEFAULT_WINDOW_SIZE)
        self.setStyleSheet(f"""
            QMainWindow {{ background-color: {UI_COLORS["background"]}; }}
            QSplitter::handle {{ background-color: {UI_COLORS["border"]}; width: 3px; height: 3px; }}
            QSplitter::handle:hover {{ background-color: #adb5bd; }}
            QLabel {{ font-weight: bold; color: {UI_COLORS["text"]}; padding: 5px; background-color: #e9ecef; border-bottom: 1px solid {UI_COLORS["border"]}; }}
            QWidget {{ font-family: "Microsoft YaHei", "Arial", sans-serif; color: {UI_COLORS["text"]}; background-color: {UI_COLORS["white"]}; }}
            QPushButton {{ background-color: {UI_COLORS["white"]}; border: 1px solid #ced4da; border-radius: 4px; padding: 8px 15px; min-height: 20px; color: {UI_COLORS["text_secondary"]}; }}
            QPushButton:hover {{ background-color: {UI_COLORS["background"]}; border-color: #6c757d; color: {UI_COLORS["text"]}; }}
            QPushButton:pressed {{ background-color: #e9ecef; }}
            QPushButton:disabled {{ background-color: #e9ecef; color: #6c757d; border-color: {UI_COLORS["border"]}; }}
            QComboBox {{ background-color: {UI_COLORS["white"]}; border: 1px solid #ced4da; border-radius: 3px; padding: 4px 8px; color: {UI_COLORS["text_secondary"]}; }}
            QComboBox:hover {{ border-color: #6c757d; }}
            QCheckBox {{ color: {UI_COLORS["text_secondary"]}; }}
            QSlider::groove:horizontal {{ background-color: {UI_COLORS["border"]}; height: 8px; border-radius: 4px; }}
            QSlider::handle:horizontal {{ background-color: #6c757d; border: 1px solid {UI_COLORS["text_secondary"]}; width: 18px; border-radius: 9px; margin: -5px 0; }}
            QSlider::handle:horizontal:hover {{ background-color: {UI_COLORS["text_secondary"]}; }}
        """)
        
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QHBoxLayout(central_widget)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # 创建分割器
        main_splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(main_splitter)
        left_splitter = QSplitter(Qt.Vertical)
        
        # 创建图形视图面板
        graphics_panel = QWidget()
        graphics_layout = QVBoxLayout(graphics_panel)
        graphics_layout.setContentsMargins(0, 0, 0, 0)
        graphics_layout.setSpacing(0)
        
        # 添加图形视图标题
        graphics_title = QLabel("图纸视图 & OCR识别")
        graphics_title.setStyleSheet(f"QLabel {{ background-color: {UI_COLORS['primary']}; color: white; font-weight: bold; padding: 8px; margin: 0px; border: none; }}")
        graphics_layout.addWidget(graphics_title)
        
        # 设置OCR面板
        self.setup_compact_ocr_panel(graphics_layout)
        
        # 创建和配置GraphicsView
        logger.debug("创建GraphicsView实例")
        self.graphics_view = GraphicsView()
        self.graphics_view.setScene(self.graphics_scene)
        graphics_layout.addWidget(self.graphics_view)
        
        # 添加标注面板
        annotation_panel = QWidget()
        annotation_layout = QVBoxLayout(annotation_panel)
        annotation_layout.setContentsMargins(0, 0, 0, 0)
        annotation_layout.setSpacing(0)
        
        # 标注面板标题
        annotation_title = QLabel("标注列表")
        annotation_title.setStyleSheet(f"QLabel {{ background-color: {UI_COLORS['secondary']}; color: white; font-weight: bold; padding: 8px; margin: 0px; border: none; }}")
        annotation_layout.addWidget(annotation_title)
        
        # 创建标注表格
        self.annotation_table = AnnotationTable()
        annotation_layout.addWidget(self.annotation_table)
        
        # 组织左侧面板
        left_splitter.addWidget(graphics_panel)
        left_splitter.addWidget(annotation_panel)
        left_splitter.setStretchFactor(0, 3)
        left_splitter.setStretchFactor(1, 1)
        
        # 创建右侧面板
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(0)
        
        # 右侧面板标题
        property_title = QLabel("属性编辑器")
        property_title.setStyleSheet(f"QLabel {{ background-color: {UI_COLORS['success']}; color: white; font-weight: bold; padding: 8px; margin: 0px; border: none; }}")
        right_layout.addWidget(property_title)
        
        # 创建属性编辑器
        self.property_editor = PropertyEditor(self)
        right_layout.addWidget(self.property_editor)
        
        # 组织主分割器
        main_splitter.addWidget(left_splitter)
        main_splitter.addWidget(right_panel)
        main_splitter.setStretchFactor(0, 3)
        main_splitter.setStretchFactor(1, 1)

        # 创建状态栏
        self.status_bar = self.statusBar()
        self.status_bar.showMessage("就绪", 2000)
        
        # 为状态栏创建PDF导航控件
        self.setup_pdf_navigation_controls()
        
        # 设置菜单栏和工具栏
        self.setup_menu_bar()
        self.setup_toolbar()
        self.setup_connections()
        
        # 设置加载对话框的大小
        if hasattr(self, 'loading_dialog'):
            self.loading_dialog.resize(self.size())
            
        # 确保创建完成后进行事件处理
        QApplication.processEvents()
        logger.debug("UI初始化完成")
        
    def setup_pdf_navigation_controls(self):
        """设置PDF导航控件（放在状态栏右侧）"""
        # 创建一个小部件来容纳导航控件
        self.pdf_nav_widget = QWidget()
        self.pdf_nav_layout = QHBoxLayout(self.pdf_nav_widget)
        self.pdf_nav_layout.setContentsMargins(0, 0, 5, 0)
        self.pdf_nav_layout.setSpacing(5)
        
        # 创建导航按钮和标签
        self.prev_page_btn = QPushButton("◀ 上一页")
        self.prev_page_btn.setMaximumWidth(80)
        self.prev_page_btn.setToolTip("显示上一页 (快捷键: 左方向键)")
        self.prev_page_btn.setEnabled(False)
        self.prev_page_btn.clicked.connect(self.go_to_prev_page)
        self.pdf_nav_layout.addWidget(self.prev_page_btn)
        
        self.page_label = QLabel("页码: 0 / 0")
        self.page_label.setFixedWidth(80)
        self.page_label.setAlignment(Qt.AlignCenter)
        self.pdf_nav_layout.addWidget(self.page_label)
        
        self.next_page_btn = QPushButton("下一页 ▶")
        self.next_page_btn.setMaximumWidth(80)
        self.next_page_btn.setToolTip("显示下一页 (快捷键: 右方向键)")
        self.next_page_btn.setEnabled(False)
        self.next_page_btn.clicked.connect(self.go_to_next_page)
        self.pdf_nav_layout.addWidget(self.next_page_btn)
        
        self.go_to_page_btn = QPushButton("前往...")
        self.go_to_page_btn.setMaximumWidth(60)
        self.go_to_page_btn.setToolTip("跳转到指定页面")
        self.go_to_page_btn.setEnabled(False)
        self.go_to_page_btn.clicked.connect(self.show_go_to_page_dialog)
        self.pdf_nav_layout.addWidget(self.go_to_page_btn)
        
        # 将导航小部件添加到状态栏右侧
        self.status_bar.addPermanentWidget(self.pdf_nav_widget)
        
        # 默认隐藏导航控件
        self.pdf_nav_widget.setVisible(False)
        
    def setup_menu_bar(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("文件")
        open_action = QAction("打开...", self); open_action.setShortcut("Ctrl+O"); open_action.triggered.connect(self.open_file); file_menu.addAction(open_action)
        
        # 添加PDF转换功能
        convert_pdf_action = QAction("PDF转换为PNG...", self); convert_pdf_action.triggered.connect(self.convert_pdf_to_images); file_menu.addAction(convert_pdf_action)
        
        if HAS_EXCEL_SUPPORT:
            export_action = QAction("导出为Excel...", self); export_action.setShortcut("Ctrl+E"); export_action.triggered.connect(self.export_to_excel); file_menu.addAction(export_action)
            export_template_action = QAction("导出到检验报告模板...", self); export_template_action.setShortcut("Ctrl+T"); export_template_action.triggered.connect(self.export_to_template); file_menu.addAction(export_template_action)
        file_menu.addSeparator()
        
        # --- 新增：创建全局快捷键动作 ---
        self.audit_action = QAction("审核", self)
        self.audit_action.setShortcut(QKeySequence("F2"))
        self.audit_action.triggered.connect(self.audit_current_annotation)
        self.addAction(self.audit_action) # 添加到主窗口，使其全局有效
        # -----------------------------

        exit_action = QAction("退出", self); exit_action.setShortcut("Ctrl+Q"); exit_action.triggered.connect(self.close); file_menu.addAction(exit_action)
    
    def setup_toolbar(self):
        toolbar = self.addToolBar("主工具栏")
        open_action = QAction("打开文件", self); open_action.triggered.connect(self.open_file); toolbar.addAction(open_action)
        if HAS_EXCEL_SUPPORT:
            export_btn = QPushButton("导出Excel"); export_btn.setToolTip("将当前标注列表导出为Excel文件"); export_btn.clicked.connect(self.export_to_excel); toolbar.addWidget(export_btn)
            export_template_btn = QPushButton("导出检验报告"); export_template_btn.setToolTip("将当前标注列表导出到检验报告模板"); export_template_btn.clicked.connect(self.export_to_template); toolbar.addWidget(export_template_btn)
        
        toolbar.addSeparator()
        toolbar.addWidget(QLabel("PDF质量:")); self.pdf_quality_combo = QComboBox(); self.pdf_quality_combo.addItems(list(PDF_QUALITY_OPTIONS.keys())); self.pdf_quality_combo.setCurrentText("高清 (4x)"); self.pdf_quality_combo.setToolTip("渲染PDF时的清晰度，越高越清晰但加载越慢"); toolbar.addWidget(self.pdf_quality_combo)
        
        # 添加强制分辨率复选框
        self.force_resolution_checkbox = QCheckBox("强制原始分辨率")
        self.force_resolution_checkbox.setToolTip("勾选后，将严格使用选定的PDF质量，即使可能导致内存占用过高。\n适合内存充足且需要最高清晰度的场景。")
        toolbar.addWidget(self.force_resolution_checkbox)
        
        toolbar.addSeparator()
        ai_recognize_action = QAction("AI识别", self); ai_recognize_action.triggered.connect(self.simulate_ai_recognition); toolbar.addAction(ai_recognize_action)
        self.area_select_action = QAction("区域OCR标注", self); self.area_select_action.setCheckable(True); self.area_select_action.setShortcut("Q"); self.area_select_action.setStatusTip("激活后，在图纸上拖拽鼠标以框选区域进行OCR识别"); self.area_select_action.toggled.connect(self.toggle_area_selection); toolbar.addAction(self.area_select_action)
        self.mask_select_action = QAction("🚫 屏蔽区域", self); self.mask_select_action.setCheckable(True); self.mask_select_action.setStatusTip("激活后，在图纸上拖拽鼠标以选择要忽略OCR的区域"); self.mask_select_action.toggled.connect(self.toggle_mask_selection); toolbar.addAction(self.mask_select_action)
        toolbar.addSeparator()
        # 添加重新排序按钮
        reorder_action = QAction("🔄 重新排序", self)
        reorder_action.setToolTip("重新给所有气泡标注排序编号(从左到右，从上到下)")
        reorder_action.triggered.connect(self.reorder_annotations)
        toolbar.addAction(reorder_action)
        clear_action = QAction("清除标注", self); clear_action.triggered.connect(self.clear_annotations); toolbar.addAction(clear_action)
        toolbar.addSeparator()
        toolbar.addWidget(QLabel("气泡大小:")); 
        # 使用常量定义比例滑块参数
        self.size_slider = QSlider(Qt.Horizontal)
        # 使用常量定义，但强制转换为整数类型
        self.size_slider.setRange(int(BUBBLE_SIZE_MIN_PERCENT), int(BUBBLE_SIZE_MAX_PERCENT))
        self.size_slider.setSingleStep(int(BUBBLE_SIZE_STEP))
        self.size_slider.setPageStep(int(BUBBLE_SIZE_STEP * 2))
        self.size_slider.setTickPosition(QSlider.TicksBelow)
        self.size_slider.setTickInterval(int(BUBBLE_SIZE_STEP*2))
        # 设置默认值
        self.size_slider.setValue(int(BUBBLE_SIZE_DEFAULT_PERCENT))
        # 调试信息
        print(f"滑块初始设置: 范围{BUBBLE_SIZE_MIN_PERCENT}-{BUBBLE_SIZE_MAX_PERCENT}, 步长{BUBBLE_SIZE_STEP}, 当前值{BUBBLE_SIZE_DEFAULT_PERCENT}")
        self.size_slider.setFixedWidth(120)  # 增加宽度，使滑块更容易拖动
        
        # 创建百分比输入框代替原来的标签
        self.size_input = QLineEdit(f"{BUBBLE_SIZE_DEFAULT_PERCENT}")
        self.size_input.setFixedWidth(40)
        self.size_input.setAlignment(Qt.AlignCenter)
        # 设置输入验证器，只允许输入数字
        self.size_input.setValidator(QIntValidator(int(BUBBLE_SIZE_MIN_PERCENT), int(BUBBLE_SIZE_MAX_PERCENT)))
        # 添加百分比符号标签
        self.percent_label = QLabel("%")
        
        # 保留标签用于显示百分比（当使用滑块时）
        self.size_label = QLabel(f"{BUBBLE_SIZE_DEFAULT_PERCENT}%")
        self.size_label.setFixedWidth(40)
        self.size_label.setVisible(False)  # 默认隐藏，因为我们现在使用输入框
        
        toolbar.addWidget(self.size_slider)
        toolbar.addWidget(self.size_input)
        toolbar.addWidget(self.percent_label)
        toolbar.addWidget(self.size_label)  # 仍然添加但隐藏
        toolbar.addSeparator()
        self.color_button = QPushButton("颜色"); self.color_button.setToolTip("选择下一个标注的颜色，或修改当前选中标注的颜色"); self.color_button.clicked.connect(self.select_annotation_color); toolbar.addWidget(self.color_button)
        toolbar.addWidget(QLabel("形状:")); self.shape_combo = QComboBox(); self.shape_combo.addItems(["空心圆", "实心圆", "五角星", "三角形"]); toolbar.addWidget(self.shape_combo)
        toolbar.addWidget(QLabel("快速样式:")); self.style_combo = QComboBox(); self.style_combo.addItems(["自定义"] + list(STYLE_NAME_MAP.values())); toolbar.addWidget(self.style_combo)
    
    def setup_connections(self):
        self.annotation_table.annotation_selected.connect(self.select_annotation_by_id)
        self.graphics_view.area_selected.connect(self.handle_area_selection)
        self.size_slider.valueChanged.connect(self.change_annotation_size)
        self.size_slider.valueChanged.connect(self.sync_size_input_from_slider)  # 同步滑块值到输入框
        self.size_input.editingFinished.connect(self.on_size_input_changed)  # 输入框编辑完成时更新气泡大小
        self.shape_combo.currentTextChanged.connect(self.change_current_annotation_shape)
        self.style_combo.currentTextChanged.connect(self.change_current_annotation_style)
        self.confidence_slider.valueChanged.connect(lambda v: self.confidence_label.setText(f"{v/100:.2f}"))
        self.ocr_button.clicked.connect(self.start_ocr_recognition)
        self.create_all_btn.clicked.connect(self.create_annotations_from_ocr)
        self.clear_ocr_btn.clicked.connect(self.clear_ocr_results)
        self.filter_combo.currentTextChanged.connect(self.filter_ocr_results)
        # --- 新增连接：连接新按钮的信号 ---
        self.property_editor.audit_requested.connect(self.audit_current_annotation)
        self.property_editor.delete_requested.connect(self.delete_current_annotation)
        
        # GPU和CPU选项互斥
        self.gpu_checkbox.toggled.connect(self.on_gpu_checkbox_toggled)
        self.cpu_checkbox.toggled.connect(self.on_cpu_checkbox_toggled)
        
        # 添加PDF导航快捷键
        self.left_action = QAction("左方向键", self)
        self.left_action.setShortcut("Left")
        self.left_action.triggered.connect(self.go_to_prev_page)
        self.addAction(self.left_action)
        
        self.right_action = QAction("右方向键", self)
        self.right_action.setShortcut("Right")
        self.right_action.triggered.connect(self.go_to_next_page)
        self.addAction(self.right_action)

    def keyPressEvent(self, event):
        """处理键盘事件"""
        # 处理方向键事件（如果有PDF打开）
        if self.pdf_file_path:
            if event.key() == Qt.Key_Left:
                self.go_to_prev_page()
                return
            elif event.key() == Qt.Key_Right:
                self.go_to_next_page()
                return
        
        super().keyPressEvent(event)

    def audit_current_annotation(self):
        if not self.current_annotation:
            # 如果没有选中的，尝试选中第一个未审核的
            sorted_annotations = sorted(self.annotations, key=lambda ann: ann.annotation_id)
            first_unaudited = next((ann for ann in sorted_annotations if not ann.is_audited), None)
            if first_unaudited:
                self.select_annotation_by_id(first_unaudited.annotation_id)
            else:
                QMessageBox.warning(self, "提示", "没有需要审核的标注。")
            return

        # 1. 审核当前项
        self.current_annotation.set_audited(True)

        # 2. 寻找下一个未审核项
        sorted_annotations = sorted(self.annotations, key=lambda ann: ann.annotation_id)
        current_index = -1
        for i, ann in enumerate(sorted_annotations):
            if ann.annotation_id == self.current_annotation.annotation_id:
                current_index = i
                break

        next_annotation_to_select = None
        
        # 从当前项之后开始寻找
        if current_index != -1:
            for i in range(current_index + 1, len(sorted_annotations)):
                if not sorted_annotations[i].is_audited:
                    next_annotation_to_select = sorted_annotations[i]
                    break
        
        # 如果后面没有，就从头开始找
        if not next_annotation_to_select:
            for ann in sorted_annotations:
                if not ann.is_audited:
                    next_annotation_to_select = ann
                    break
        
        # 3. 跳转
        if next_annotation_to_select:
            self.select_annotation_by_id(next_annotation_to_select.annotation_id)
        else:
            QMessageBox.information(self, "审核完成", "所有标注均已审核！")

    def export_to_excel(self):
        if not HAS_EXCEL_SUPPORT:
            QMessageBox.warning(self, "功能缺失", "缺少 'openpyxl' 库，无法导出Excel。\n请运行: pip install openpyxl"); return
        if not self.annotations:
            QMessageBox.information(self, "提示", "标注列表为空，无需导出。"); return
        
        default_filename = f"{Path(self.current_file_path).stem}_标注列表.xlsx" if self.current_file_path else "标注列表.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "导出为Excel文件", default_filename, "Excel 文件 (*.xlsx)")
        
        if not file_path: return
        
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "标注数据"
            headers = [self.annotation_table.horizontalHeaderItem(i).text() for i in range(self.annotation_table.columnCount())]
            ws.append(headers)
            
            header_font = Font(bold=True)
            for cell in ws[1]: 
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for ann in self.annotations:
                row_data = [
                    str(ann.annotation_id),
                    ann.dimension_type,
                    ann.dimension,
                    ann.upper_tolerance,
                    ann.lower_tolerance,
                    "是" if ann.is_audited else "否"
                ]
                ws.append(row_data)

            for col_idx, column in enumerate(ws.columns):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max(max_length + 2, len(headers[col_idx]) + 2)
                ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50
            ws.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10


            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"标注列表已成功导出到:\n{file_path}")
            self.status_bar.showMessage(f"成功导出到 {Path(file_path).name}", 5000)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出到Excel时发生错误:\n{e}")
            self.status_bar.showMessage("导出失败", 3000)
            
    def export_to_template(self):
        """将标注列表导出到Excel模板中
        - 将序号、类型、尺寸、上公差、下公差插入到A-E列
        - 从第14行开始插入
        """
        try:
            import xlwings as xw
            HAS_XLWINGS_SUPPORT = True
        except ImportError:
            HAS_XLWINGS_SUPPORT = False
            if not HAS_EXCEL_SUPPORT:
                QMessageBox.warning(self, "功能缺失", "缺少Excel支持库。\n请运行: pip install xlwings 或 pip install openpyxl")
                return
        
        if not self.annotations:
            QMessageBox.information(self, "提示", "标注列表为空，无需导出。")
            return
        
        # 默认使用根目录下的muban.xlsx
        template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "muban.xlsx")
        if not os.path.exists(template_path):
            # 如果默认模板不存在，则提示选择
            template_path, _ = QFileDialog.getOpenFileName(self, "选择Excel模板文件", "", "Excel 文件 (*.xlsx)")
            if not template_path:
                return
        
        # 选择保存位置
        default_filename = f"{Path(self.current_file_path).stem}_检验报告.xlsx" if self.current_file_path else "检验报告.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", default_filename, "Excel 文件 (*.xlsx)")
        if not save_path:
            return
        
        try:
            # 获取要插入的标注数据
            annotations_data = []
            for ann in self.annotations:
                row_data = [
                    str(ann.annotation_id),
                    ann.dimension_type,
                    ann.dimension,
                    ann.upper_tolerance,
                    ann.lower_tolerance
                ]
                annotations_data.append(row_data)
            
            # 排序标注（按ID排序）
            annotations_data.sort(key=lambda x: int(x[0]) if x[0].isdigit() else float('inf'))
            
            # 确定插入行的范围
            start_row = 14  # 从第14行开始
            insert_count = len(annotations_data)  # 需要插入的行数
            
            if HAS_XLWINGS_SUPPORT:
                # 使用xlwings插入行 (这种方式会更接近Excel手动操作)
                try:
                    # 先复制模板到保存位置
                    import shutil
                    shutil.copy2(template_path, save_path)
                    
                    # 用xlwings打开文件
                    app = xw.App(visible=False)
                    wb = app.books.open(save_path)
                    ws = wb.sheets[0]
                    
                    # 插入行 - 这会像Excel手动操作一样插入干净的行
                    # 注意xlwings中行号从1开始计数
                    # 限制只在A-P列插入，不影响Q-S列
                    ws.range(f"A{start_row}:P{start_row+insert_count-1}").insert('down')
                    
                    # 填充数据 (可选)
                    for i, row_data in enumerate(annotations_data):
                        row_idx = start_row + i
                        # 只处理A-E列
                        for j, value in enumerate(row_data):
                            col_letter = chr(65 + j)  # A=65, B=66, ...
                            ws.range(f"{col_letter}{row_idx}").value = value
                        
                        # 计算并填充O列和P列
                        # O列=C+D（尺寸+上公差）
                        dimension = row_data[2] if len(row_data) > 2 else ""
                        upper_tol = row_data[3] if len(row_data) > 3 else ""
                        lower_tol = row_data[4] if len(row_data) > 4 else ""
                        
                        # 尝试进行数值计算，如果是数字的话
                        # O列=C+D（尺寸+上公差）
                        try:
                            # 尝试将尺寸和公差转换为浮点数进行计算
                            if dimension and upper_tol:
                                try:
                                    dim_value = float(dimension)
                                    # 处理公差值，去掉前面的+号
                                    tol_value = float(upper_tol.replace('+', '')) if upper_tol.startswith('+') else float(upper_tol)
                                    # 计算结果
                                    result_value = dim_value + tol_value
                                    # 设置单元格值为计算结果
                                    ws.range(f"O{row_idx}").value = result_value
                                except ValueError:
                                    # 如果无法转换为数值，则留空
                                    ws.range(f"O{row_idx}").value = ""
                            else:
                                ws.range(f"O{row_idx}").value = ""
                        except Exception as e:
                            print(f"计算O列时出错: {e}")
                            ws.range(f"O{row_idx}").value = ""
                        
                        # P列=C+E（尺寸+下公差）
                        try:
                            # 尝试将尺寸和公差转换为浮点数进行计算
                            if dimension and lower_tol:
                                try:
                                    dim_value = float(dimension)
                                    # 处理公差值，去掉前面的+号
                                    tol_value = float(lower_tol.replace('+', '')) if lower_tol.startswith('+') else float(lower_tol)
                                    # 计算结果
                                    result_value = dim_value + tol_value
                                    # 设置单元格值为计算结果
                                    ws.range(f"P{row_idx}").value = result_value
                                except ValueError:
                                    # 如果无法转换为数值，则留空
                                    ws.range(f"P{row_idx}").value = ""
                            else:
                                ws.range(f"P{row_idx}").value = ""
                        except Exception as e:
                            print(f"计算P列时出错: {e}")
                            ws.range(f"P{row_idx}").value = ""
                    
                    # 保存文件并关闭Excel
                    wb.save()
                    wb.close()
                    app.quit()
                    
                    QMessageBox.information(self, "导出成功", f"标注列表已成功导出到:\n{save_path}")
                    self.status_bar.showMessage(f"成功导出到 {Path(save_path).name}", 5000)
                    return
                except Exception as e:
                    # 如果xlwings出错，回退到openpyxl
                    QMessageBox.warning(self, "提示", f"使用xlwings导出失败 ({str(e)})，将尝试使用openpyxl。")
            
            # 如果没有xlwings支持或xlwings失败，使用openpyxl
            if HAS_EXCEL_SUPPORT:
                # 打开模板文件
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
                
                # 保存Q13-S30区域的内容
                q_s_content = {}
                for r in range(start_row, start_row + insert_count + 30):  # 保存足够多的行
                    for c in range(17, 20):  # Q=17, R=18, S=19
                        try:
                            cell_coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
                            q_s_content[cell_coord] = ws[cell_coord].value
                        except:
                            continue
                
                # 插入行
                ws.insert_rows(start_row, insert_count)
                
                # 检查并手动取消新插入行中的合并单元格
                for r in range(start_row, start_row + insert_count):
                    # 检查每个单元格是否是合并单元格的一部分
                    for c in range(1, 17):  # A列到P列
                        try:
                            # 获取单元格坐标
                            coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
                            
                            # 检查该单元格是否是合并单元格的一部分
                            for merged_range in list(ws.merged_cells.ranges):
                                if coord in merged_range:
                                    # 如果是合并单元格，解除合并
                                    ws.unmerge_cells(str(merged_range))
                                    break
                        except:
                            continue
                
                # 恢复Q-S列的内容
                for cell_coord, value in q_s_content.items():
                    ws[cell_coord] = value
                
                # 填充数据 (可选)
                for i, row_data in enumerate(annotations_data):
                    row_idx = start_row + i
                    # 只处理A-E列的数据
                    for j, value in enumerate(row_data):
                        if j < len(row_data):  # 确保不越界
                            ws.cell(row=row_idx, column=j+1).value = value
                    
                    # 计算并填充O列和P列
                    # O列=C+D（尺寸+上公差）
                    dimension = row_data[2] if len(row_data) > 2 else ""
                    upper_tol = row_data[3] if len(row_data) > 3 else ""
                    lower_tol = row_data[4] if len(row_data) > 4 else ""
                    
                    # 尝试进行数值计算，如果是数字的话
                    # O列=C+D（尺寸+上公差）
                    try:
                        # 尝试将尺寸和公差转换为浮点数进行计算
                        if dimension and upper_tol:
                            try:
                                dim_value = float(dimension)
                                # 处理公差值，去掉前面的+号
                                tol_value = float(upper_tol.replace('+', '')) if upper_tol.startswith('+') else float(upper_tol)
                                # 计算结果
                                result_value = dim_value + tol_value
                                # 设置单元格值为计算结果
                                ws.cell(row=row_idx, column=15).value = result_value  # O列是第15列
                            except ValueError:
                                # 如果无法转换为数值，则留空
                                ws.cell(row=row_idx, column=15).value = ""
                        else:
                            ws.cell(row=row_idx, column=15).value = ""
                    except Exception as e:
                        print(f"计算O列时出错: {e}")
                        ws.cell(row=row_idx, column=15).value = ""
                    
                    # P列=C+E（尺寸+下公差）
                    try:
                        # 尝试将尺寸和公差转换为浮点数进行计算
                        if dimension and lower_tol:
                            try:
                                dim_value = float(dimension)
                                # 处理公差值，去掉前面的+号
                                tol_value = float(lower_tol.replace('+', '')) if lower_tol.startswith('+') else float(lower_tol)
                                # 计算结果
                                result_value = dim_value + tol_value
                                # 设置单元格值为计算结果
                                ws.cell(row=row_idx, column=16).value = result_value  # P列是第16列
                            except ValueError:
                                # 如果无法转换为数值，则留空
                                ws.cell(row=row_idx, column=16).value = ""
                        else:
                            ws.cell(row=row_idx, column=16).value = ""
                    except Exception as e:
                        print(f"计算P列时出错: {e}")
                        ws.cell(row=row_idx, column=16).value = ""
                
                # 保存文件
                wb.save(save_path)
                QMessageBox.information(self, "导出成功", f"标注列表已成功导出到:\n{save_path}")
                self.status_bar.showMessage(f"成功导出到 {Path(save_path).name}", 5000)
            else:
                QMessageBox.critical(self, "导出失败", "未找到可用的Excel处理库")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出到Excel模板时发生错误:\n{e}")
            self.status_bar.showMessage("导出失败", 3000)

    def open_file(self):
        """打开文件对话框选择文件"""
        # 创建文件对话框并设置过滤器
        file_dialog = QFileDialog(self)
        file_dialog.setWindowTitle("打开文件")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter(FILE_DIALOG_FILTER)
        
        # 设置默认目录
        if self.current_file_path:
            dir_path = str(Path(self.current_file_path).parent)
            file_dialog.setDirectory(dir_path)
        
        # 显示对话框
        if file_dialog.exec():
            file_paths = file_dialog.selectedFiles()
            if file_paths:
                self.status_bar.showMessage(f"正在加载文件: {file_paths[0]}...")
                self.load_file(file_paths[0])

    def load_file(self, file_path: str):
        """加载文件
        
        Args:
            file_path: 文件路径
        """
        if not file_path or not Path(file_path).exists():
            QMessageBox.warning(self, "错误", f"文件不存在: {file_path}")
            self.status_bar.clearMessage()
            return
            
        # 清除当前选中的标注
        if self.current_annotation:
            self.current_annotation = None
            self.property_editor.set_annotation(None, None, None)
        
        # 转换为Path对象以便于处理
        file_path = Path(file_path)
        extension = file_path.suffix.lower()
        
        # 确保扩展名有效
        if extension not in SUPPORTED_IMAGE_FORMATS + SUPPORTED_PDF_FORMATS + SUPPORTED_DXF_FORMATS:
            QMessageBox.warning(self, "错误", f"不支持的文件格式: {extension}")
            self.status_bar.clearMessage()
            return
        
        # 显示加载对话框
        self.loading_label.setText(f"⏳ 正在加载文件...\n{file_path.name}")
        self.loading_dialog.resize(self.size())
        self.loading_dialog.show()
        QApplication.processEvents()  # 确保UI立即更新
        
        # 清除当前状态
        self.clear_annotations(show_empty_message=False)
        self.annotation_table.clear_annotations()  # 确保表格也被清空
        self.clear_ocr_results()
        self.graphics_scene.clear()
        self.annotation_counter = 0  # 重置标注计数器
        
        # 重置PDF相关状态
        self.pdf_file_path = None
        self.pdf_page_count = 0
        self.current_pdf_page = 0
        self.pdf_pages_cache.clear()
        self.annotations_by_page.clear()
        self.ocr_results_by_page.clear()
        
        # 隐藏PDF导航控件
        self.pdf_nav_widget.setVisible(False)
        
        # 创建并启动文件加载线程
        file_loader = FileLoaderWorker(
            str(file_path),
            pdf_quality=self.pdf_quality_combo.currentText()
        )
        
        # 连接信号
        file_loader.signals.progress.connect(self._on_file_load_progress)
        file_loader.signals.finished.connect(self._on_file_loaded)
        file_loader.signals.pdf_loaded.connect(self._on_pdf_file_loaded)
        file_loader.signals.error.connect(self._on_file_load_error)
        
        # 启动线程
        self.thread_pool.start(file_loader)

    def _on_file_load_progress(self, progress: int, message: str):
        """文件加载进度更新"""
        self.loading_label.setText(f"⏳ {message}\n({progress}%)")
        self.status_bar.showMessage(message)
        QApplication.processEvents()

    def _on_file_loaded(self, file_path: str, pixmap: QPixmap):
        """文件加载完成回调
        
        Args:
            file_path: 文件路径
            pixmap: 图像数据
        """
        try:
            self.current_file_path = file_path
            self.current_pixmap = pixmap
            
            # 清除当前场景
            self.graphics_scene.clear()
            
            # 添加图像到场景
            self.graphics_scene.addPixmap(pixmap)
            logger.debug(f"图像已添加到场景，尺寸: {pixmap.width()}x{pixmap.height()}")
            
            # 初始化标注ID计数器
            self.annotation_counter = 0
            
            # 执行一次居中操作，之后不干扰用户缩放
            self.center_view()
            
            # 更新状态栏
            file_name = Path(file_path).name
            self.status_bar.showMessage(f"已加载: {file_name} ({pixmap.width()}x{pixmap.height()})", 5000)
            
            # 隐藏加载对话框
            self.loading_dialog.hide()
        except Exception as e:
            self._on_file_load_error(f"处理加载结果时出错: {str(e)}")
            logger.exception("图像加载完成处理错误")

    def _on_pdf_file_loaded(self, file_path: str, page_count: int):
        """PDF文件加载完成"""
        try:
            # 设置PDF相关属性
            self.pdf_file_path = file_path
            self.pdf_page_count = page_count
            self.current_pdf_page = 0  # 从第一页开始
            
            # 如果是多页PDF，显示导航控件
            if page_count > 1:
                self.update_pdf_navigation_controls()
                
                # 在状态栏显示提示消息，而不是弹窗
                self.status_bar.showMessage(f"多页PDF文件，共 {page_count} 页，使用右下角导航控件或键盘方向键切换页面", 5000)
            
            # 加载第一页
            self.load_pdf_page(0)
            
        except Exception as e:
            self._on_file_load_error(f"处理PDF文件时出错: {str(e)}")

    def _on_file_load_error(self, error_msg: str):
        """文件加载错误处理"""
        QMessageBox.warning(self, "加载错误", error_msg)
        self.status_bar.showMessage(f"❌ 文件加载失败: {error_msg}", 5000)
        self.loading_dialog.hide()

    def simulate_ai_recognition(self):
        self.start_ocr_recognition()

    def start_ocr_recognition(self):
        """启动OCR识别过程"""
        if not HAS_OCR_SUPPORT:
            QMessageBox.warning(self, "功能缺失", "OCR功能需要PaddleOCR和依赖包。请安装所需依赖。")
            return
        
        if not self.current_pixmap:
            QMessageBox.information(self, "提示", "请先打开图片文件。")
            return
            
        # 保存已有的OCR结果，而不是清除
        existing_results = self.ocr_results.copy()
        # 只清除显示，不清除结果数据
        self.clear_ocr_display()
        
        # 获取语言配置
        lang_text = self.language_combo.currentText()
        lang_code = DEFAULT_OCR_LANGUAGES.get(lang_text, ["ch_sim"])
        
        # 获取环境配置
        force_cpu = self.cpu_checkbox.isChecked()
        use_gpu = self.gpu_checkbox.isChecked() and not force_cpu
        
        # 获取CPU线程数
        cpu_threads = self.threads_spinbox.value()
        
        # 显示设备模式
        device_mode = "CPU" if force_cpu else ("GPU" if use_gpu else "自动")
        self.status_bar.showMessage(f"正在进行OCR文本识别... (使用{device_mode}模式)")
        
        # 获取屏蔽区域数据
        masked_regions_data = [{'x': r.x(), 'y': r.y(), 'width': r.width(), 'height': r.height()} for r in self.masked_regions]
        
        # 创建OCR工作器
        self.ocr_worker = PaddleOCRWorker(
            self.current_file_path, 
            lang_code, 
            masked_regions_data,
            force_cpu=force_cpu,
            cpu_threads=cpu_threads,  # 传递线程数
            direct_recognition=False  # 全图OCR不使用直接识别模式
        )
        
        # 连接信号
        self.ocr_worker.signals.progress.connect(self.on_ocr_progress)
        self.ocr_worker.signals.error.connect(self.on_ocr_error)
        
        # 修改on_ocr_finished连接，合并现有结果
        self.ocr_worker.signals.finished.connect(
            lambda results: self.on_ocr_finished(results, existing_results)
        )
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 启动线程
        self.thread_pool.start(self.ocr_worker)

    def on_ocr_progress(self, progress: int):
        self.progress_bar.setValue(progress)

    def on_ocr_error(self, error_msg: str):
        self.ocr_button.setEnabled(True); self.ocr_button.setText("🔍 开始OCR识别")
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "OCR识别错误", error_msg)

    def merge_adjacent_ocr_results(self, results: List[dict]) -> List[dict]:
        """合并可能属于同一尺寸标注的相邻OCR结果"""
        if not results or len(results) <= 1:
            return results
            
        # 按照y坐标排序结果，对于相同y坐标的，按照x坐标排序
        # 这样可以确保同一行的文本从左到右处理
        sorted_results = sorted(results, key=lambda r: (r.get('center_y', 0), r.get('center_x', 0)))
        print(f"📊 排序后的OCR结果: {[(i, r.get('text', ''), r.get('center_x', 0), r.get('center_y', 0)) for i, r in enumerate(sorted_results)]}")
        
        # 检测所有可能是竖排文本的结果
        vertical_texts = []
        for i, result in enumerate(sorted_results):
            # 检查边界框的高宽比，如果高大于宽很多，可能是竖排文本
            bbox_width = result.get('bbox_width', 0)
            bbox_height = result.get('bbox_height', 0)
            if bbox_height > bbox_width * 1.5 and bbox_height > 30:  # 高宽比大于1.5且高度超过30像素
                vertical_texts.append(i)
                print(f"  🔍 检测到竖排文本 [{i}]: '{result.get('text', '')}' (高宽比: {bbox_height/max(1, bbox_width):.1f})")
        
        # 使用更复杂的合并策略，允许多次合并
        # 首先，为每个结果分配一个组ID，初始时每个结果自己是一组
        groups = {i: [i] for i in range(len(sorted_results))}
        group_of_result = {i: i for i in range(len(sorted_results))}
        
        print("🔍 开始查找相邻文本...")
        
        # 预处理阶段：优先合并竖排公差文本
        # 竖排公差文本之间的合并是最高优先级
        print("🔍 预处理阶段：优先合并竖排公差文本...")
        
        # 如果有多个竖排文本，尝试合并它们
        if len(vertical_texts) > 1:
            for i, idx1 in enumerate(vertical_texts):
                result1 = sorted_results[idx1]
                text1 = result1.get('text', '').strip()
                
                # 只处理公差文本（+、-、0开头）
                if not (text1.startswith(('+', '-')) or text1 == '0' or (text1.startswith('0') and len(text1) > 1)):
                    continue
                    
                for j, idx2 in enumerate(vertical_texts[i+1:], i+1):
                    result2 = sorted_results[idx2]
                    text2 = result2.get('text', '').strip()
                    
                    # 只处理公差文本（+、-、0开头）
                    if not (text2.startswith(('+', '-')) or text2 == '0' or (text2.startswith('0') and len(text2) > 1)):
                        continue
                    
                    # 计算两个竖排文本之间的距离
                    dist_x = abs(result1.get('center_x', 0) - result2.get('center_x', 0))
                    dist_y = abs(result1.get('center_y', 0) - result2.get('center_y', 0))
                    
                    # 竖排文本合并也需要更严格的x轴距离判断
                    # 特别是对于可能分散的竖排文本（如"+0"和"03"）
                    if dist_x < 50 and dist_y < 150:  # 减小x轴阈值，原来是250
                        print(f"    ✅ 竖排公差文本距离符合: '{text1}' 和 '{text2}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
                        
                        # 获取当前两个结果所在的组
                        group_i = group_of_result[idx1]
                        group_j = group_of_result[idx2]
                        
                        # 如果已经在同一组，跳过
                        if group_i == group_j:
                            continue
                        
                        # 将j所在组的所有结果合并到i所在组
                        for idx in groups[group_j]:
                            groups[group_i].append(idx)
                            group_of_result[idx] = group_i
                        
                        # 清空j所在组
                        groups[group_j] = []
                        
                        print(f"    👉 合并竖排公差文本组: {group_j} -> {group_i}, 组{group_i}现在包含: {groups[group_i]}")
        
        # 首先，识别所有公差文本和数值文本
        tolerance_indices = []  # 公差文本索引
        numeric_indices = []    # 数值文本索引
        diameter_indices = []   # 直径符号文本索引
        
        for i, result in enumerate(sorted_results):
            text = result.get('text', '').strip()
            
            # 检查是否是公差文本（+、-、0、±开头）
            if (text.startswith(('+', '-', '±')) or 
                (text.startswith('0') and len(text) > 1) or
                text == '0'):
                tolerance_indices.append(i)
                print(f"  🔍 识别到公差文本 [{i}]: '{text}'")
            
            # 检查是否是直径符号文本
            elif text.startswith(('Φ', '∅', 'Ø')):
                diameter_indices.append(i)
                print(f"  🔍 识别到直径符号文本 [{i}]: '{text}'")
                
                # 如果直径符号后面有数字，也将其作为数值文本
                if re.search(r'[Φ∅Ø]\s*\d+', text):
                    numeric_indices.append(i)
                    print(f"  🔍 直径符号文本也包含数值，添加到数值文本列表")
            
            # 检查是否是纯数值文本
            elif re.match(r'^[\d\.]+$', text):
                numeric_indices.append(i)
                print(f"  🔍 识别到数值文本 [{i}]: '{text}'")
        
        # 第一阶段：处理公差文本与左侧数值的多重匹配
        # 允许一个公差文本与多个左侧数值匹配（上公差、下公差、基本值）
        print("🔍 第一阶段：处理公差文本与左侧数值的多重匹配...")
        
        # 为每个公差文本创建一个可能的匹配列表
        tolerance_matches = {}
        
        for i in tolerance_indices:
            current = sorted_results[i]
            current_text = current.get('text', '').strip()
            is_current_vertical = i in vertical_texts
            
            print(f"  检查公差文本 [{i}]: '{current_text}' {'(竖排)' if is_current_vertical else ''}")
            
            # 查找所有可能匹配的左侧数值文本
            possible_matches = []
            
            for j in numeric_indices:
                other = sorted_results[j]
                other_text = other.get('text', '').strip()
                is_other_vertical = j in vertical_texts
                
                # 计算距离，但更关注水平方向的距离
                dist_x = current.get('center_x', 0) - other.get('center_x', 0)
                dist_y = abs(current.get('center_y', 0) - other.get('center_y', 0))
                
                # 竖排文本使用更宽松的距离判断
                if is_current_vertical or is_other_vertical:
                    # 公差应该在数值右侧，但对于竖排文本，可能有不同的排列
                    # 所以不严格要求dist_x为正
                    # 修改：增加x轴距离的权重，使其更加重要
                    distance = abs(dist_x) * 2 + dist_y
                    
                    # 修改：减小x轴距离阈值，使匹配更加精确
                    # 原来是250，改为更合理的50
                    if abs(dist_x) < 50 and dist_y < 150:
                        possible_matches.append((j, distance))
                        print(f"    ✓ 竖排文本匹配: '{current_text}' 和 '{other_text}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
                else:
                    # 普通文本要求公差在数值右侧
                    if dist_x <= 0:
                        continue
                    
                    # 计算综合距离，水平距离权重更大
                    distance = dist_x + dist_y * 2
                    
                    # 垂直距离不能太大
                    if dist_y > 70:
                        continue
                    
                    # 水平距离不能太大
                    if dist_x > 200:
                        continue
                    
                    possible_matches.append((j, distance))
                    print(f"    ✓ 普通文本匹配: '{current_text}' 和 '{other_text}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
            
            # 按距离排序
            possible_matches.sort(key=lambda x: x[1])
            
            # 最多保留前3个最近的匹配
            best_matches = [match[0] for match in possible_matches[:3]]
            
            if best_matches:
                print(f"    ✅ 公差文本 '{current_text}' 找到 {len(best_matches)} 个匹配: {[sorted_results[idx].get('text', '') for idx in best_matches]}")
                tolerance_matches[i] = best_matches
        
        # 根据匹配关系，合并公差文本和数值文本
        print("🔍 根据匹配关系合并公差文本和数值文本...")
        
        # 创建一个新的组结构，将公差文本和匹配的数值文本合并到同一组
        for tolerance_idx, match_indices in tolerance_matches.items():
            if not match_indices:
                continue
                
            # 选择第一个匹配作为主组
            main_group_id = group_of_result[match_indices[0]]
            
            # 将公差文本合并到主组
            tolerance_group_id = group_of_result[tolerance_idx]
            
            # 如果公差文本已经在主组中，跳过
            if tolerance_group_id == main_group_id:
                continue
            
            # 将公差文本所在组的所有结果合并到主组
            for idx in groups[tolerance_group_id]:
                groups[main_group_id].append(idx)
                group_of_result[idx] = main_group_id
            
            # 清空公差文本所在组
            groups[tolerance_group_id] = []
            
            print(f"    👉 合并公差组: {tolerance_group_id} -> {main_group_id}, 组{main_group_id}现在包含: {groups[main_group_id]}")
            
            # 合并其他匹配的数值文本
            for match_idx in match_indices[1:]:
                match_group_id = group_of_result[match_idx]
                
                # 如果已经在主组中，跳过
                if match_group_id == main_group_id:
                    continue
                
                # 将匹配文本所在组的所有结果合并到主组
                for idx in groups[match_group_id]:
                    groups[main_group_id].append(idx)
                    group_of_result[idx] = main_group_id
                
                # 清空匹配文本所在组
                groups[match_group_id] = []
                
                print(f"    👉 合并匹配组: {match_group_id} -> {main_group_id}, 组{main_group_id}现在包含: {groups[main_group_id]}")
        
        # 第二阶段：处理其他常规合并情况 - 仅限公差文本
        print("🔍 第二阶段：处理其他常规合并情况 - 仅限公差文本...")
        for i in tolerance_indices:  # 只处理公差文本
            current = sorted_results[i]
            if current is None:
                continue
                
            current_text = current.get('text', '').strip()
            is_current_vertical = i in vertical_texts
            
            print(f"  检查公差文本 [{i}]: '{current_text}' {'(竖排)' if is_current_vertical else ''}")
            
            # 只与数值文本或直径文本匹配
            for j in numeric_indices + diameter_indices:
                if i == j:
                    continue
                    
                other = sorted_results[j]
                if other is None:
                    continue
                    
                other_text = other.get('text', '').strip()
                is_other_vertical = j in vertical_texts
                
                # 计算中心点距离
                dist_x = abs(current.get('center_x', 0) - other.get('center_x', 0))
                dist_y = abs(current.get('center_y', 0) - other.get('center_y', 0))
                
                # 初始化合并标志
                should_merge = False
                
                # 检查文本特征
                is_diameter = other_text.startswith(('Φ', '∅', 'Ø'))
                is_numeric = re.match(r'^[\d\.]+$', other_text) is not None
                
                # 竖排文本距离判断更宽松
                if is_current_vertical or is_other_vertical:
                    # 竖排文本有特殊的空间关系，但需要更严格的x轴阈值
                    x_threshold = 50  # 减小阈值，原来是250
                    y_threshold = 150  # 保持垂直阈值不变
                    
                    # 检查竖排文本与其他文本的匹配关系
                    if dist_x < x_threshold and dist_y < y_threshold:
                        # 检查是否组成完整标注
                        if is_diameter or is_numeric:
                            should_merge = True
                            print(f"    ✅ 竖排公差文本匹配: '{current_text}' 和 '{other_text}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
                
                # 直径符号特殊处理
                elif is_diameter:
                    if dist_x < 150 and dist_y < 50:
                        should_merge = True
                        print(f"    ✅ 公差文本与直径符号配对: '{current_text}' 和 '{other_text}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
                
                # 普通数值文本
                elif is_numeric:
                    if dist_x < 100 and dist_y < 30:
                        should_merge = True
                        print(f"    ✅ 公差文本与数值匹配: '{current_text}' 和 '{other_text}' (距离: x={dist_x:.1f}, y={dist_y:.1f})")
                
                # 如果应该合并，将两个结果合并到同一组
                if should_merge:
                    # 获取当前两个结果所在的组
                    group_i = group_of_result[i]
                    group_j = group_of_result[j]
                    
                    # 如果已经在同一组，跳过
                    if group_i == group_j:
                        continue
                    
                    # 将j所在组的所有结果合并到i所在组
                    for idx in groups[group_j]:
                        groups[group_i].append(idx)
                        group_of_result[idx] = group_i
                    
                    # 清空j所在组
                    groups[group_j] = []
                    
                    print(f"    👉 合并组: {group_j} -> {group_i}, 组{group_i}现在包含: {groups[group_i]}")
        
        # 现在我们有了分组信息，处理每个非空组
        merged_results = []
        
        print("🔄 开始处理合并组...")
        for group_id, members in groups.items():
            if not members:  # 跳过空组
                continue
                
            print(f"  处理组 {group_id}: {members}")
            
            if len(members) == 1:
                # 只有一个成员，直接添加
                merged_results.append(sorted_results[members[0]])
                continue
                
            # 有多个成员，需要合并
            # 收集组内所有文本和它们的索引，按类型分类
            diameters = []  # 直径符号文本 (Φ, ∅, Ø)
            minus_symbols = []  # 负号文本 (-)
            plus_symbols = []   # 正号文本 (+)
            pm_symbols = []     # 正负号文本 (±)
            zeros = []          # 0开头的文本
            numbers = []        # 纯数字文本
            others = []         # 其他文本
            
            # 检查是否是竖排文本组
            is_vertical_group = any(idx in vertical_texts for idx in members)
            
            # 检查是否包含公差文本
            has_tolerance = False
            
            for idx in members:
                result = sorted_results[idx]
                text = result.get('text', '').strip()
                
                if text.startswith(('Φ', '∅', 'Ø')):
                    diameters.append((idx, text))
                elif text.startswith('-'):
                    minus_symbols.append((idx, text))
                    has_tolerance = True
                elif text.startswith('+'):
                    plus_symbols.append((idx, text))
                    has_tolerance = True
                elif text.startswith('±'):
                    pm_symbols.append((idx, text))
                    has_tolerance = True
                elif text == '0' or (text.startswith('0') and len(text) > 1):
                    zeros.append((idx, text))
                    has_tolerance = True
                elif re.match(r'^[\d\.]+$', text):
                    numbers.append((idx, text))
                else:
                    others.append((idx, text))
            
            # 如果组内没有公差文本，且不是单个成员，则拆分回单独的结果
            if not has_tolerance and len(members) > 1:
                print(f"    ⚠️ 组内没有公差文本，拆分回单独的结果")
                for idx in members:
                    merged_results.append(sorted_results[idx])
                continue
            
            # 按x坐标或y坐标排序各类文本
            for text_list in [diameters, minus_symbols, plus_symbols, pm_symbols, zeros, numbers, others]:
                if is_vertical_group:
                    # 竖排文本组按y坐标排序（从上到下）
                    text_list.sort(key=lambda item: sorted_results[item[0]].get('center_y', 0))
                else:
                    # 普通文本组按x坐标排序（从左到右）
                    text_list.sort(key=lambda item: sorted_results[item[0]].get('center_x', 0))
            
            # 确定合并顺序，根据文本类型安排
            # 1. 直径符号在前，后跟数字
            # 2. 公差值按照：数值 -> 0 -> ± -> + -> -
            
            # 根据内容特点调整合并顺序
            ordered_items = []
            
            # 如果有直径符号，优先放最前面
            if diameters:
                ordered_items.extend(diameters)
            
            # 基本数值放前面，但如果直径符号已经包含数字，则不需要额外添加数字
            if diameters and len(diameters) > 0:
                diameter_text = diameters[0][1]
                # 检查直径符号文本是否已包含数字
                if not re.search(r'[Φ∅Ø]\s*\d+', diameter_text) and numbers:
                    ordered_items.extend(numbers)
            else:
                # 没有直径符号，直接添加数字
                ordered_items.extend(numbers)
            
            # 竖排文本组特殊处理
            if is_vertical_group:
                # 对于竖排文本，可能需要特殊的顺序
                # 例如，对于"+0.03"这种竖排文本，可能是"03"在上，"+0"在下
                # 或者反过来，根据实际y坐标排序
                
                # 合并剩余的所有符号和数字
                remaining_items = zeros + pm_symbols + plus_symbols + minus_symbols + others
                
                # 按y坐标排序
                remaining_items.sort(key=lambda item: sorted_results[item[0]].get('center_y', 0))
                
                ordered_items.extend(remaining_items)
            else:
                # 普通文本使用标准顺序
                # 公差值顺序：0 -> ± -> + -> -
                ordered_items.extend(zeros)
                ordered_items.extend(pm_symbols)
                ordered_items.extend(plus_symbols)
                ordered_items.extend(minus_symbols)
                
                # 其他类型放最后
                ordered_items.extend(others)
            
            if not ordered_items:
                continue
                
            # 获取排序后的文本和索引
            ordered_indices = [item[0] for item in ordered_items]
            ordered_texts = [item[1] for item in ordered_items]
            
            # 特殊处理：如果直径符号后面紧跟数字，需要添加空格
            processed_texts = []
            for i, text in enumerate(ordered_texts):
                if i > 0 and (text == '0' or text.startswith('+') or text.startswith('-')) and \
                   (processed_texts[-1].endswith(tuple('0123456789'))):
                    # 在数字和公差符号之间添加空格
                    processed_texts.append(' ' + text)
                else:
                    processed_texts.append(text)
            
            # 合并文本
            merged_text = ''.join(processed_texts)
            
            print(f"    👉 合并文本: {ordered_texts} -> '{merged_text}'")
            
            # 创建合并后的结果
            base_result = sorted_results[members[0]].copy()
            base_result['text'] = merged_text
            
            # 更新边界框
            all_bbox_points = []
            for idx in members:
                result = sorted_results[idx]
                if 'bbox' in result:
                    all_bbox_points.extend(result['bbox'])
            
            if all_bbox_points:
                x_coords = [p[0] for p in all_bbox_points]
                y_coords = [p[1] for p in all_bbox_points]
                
                min_x, max_x = min(x_coords), max(x_coords)
                min_y, max_y = min(y_coords), max(y_coords)
                
                base_result['bbox'] = [
                    [min_x, min_y],
                    [max_x, min_y],
                    [max_x, max_y],
                    [min_x, max_y]
                ]
                
                # 更新边界框宽高
                base_result['bbox_width'] = max_x - min_x
                base_result['bbox_height'] = max_y - min_y
            
            # 更新中心点
            center_x_sum = sum(sorted_results[idx].get('center_x', 0) for idx in members)
            center_y_sum = sum(sorted_results[idx].get('center_y', 0) for idx in members)
            
            base_result['center_x'] = center_x_sum / len(members)
            base_result['center_y'] = center_y_sum / len(members)
            
            # 检查是否有竖排文本特征，设置标记
            if is_vertical_group:
                base_result['is_vertical'] = True
                print(f"    ✓ 标记为竖排文本组")
            
            merged_results.append(base_result)
            
        print(f"✅ 合并完成，最终结果: {[(i, r.get('text', '')) for i, r in enumerate(merged_results)]}")
        
        # 返回合并后的结果
        return merged_results

    def on_ocr_finished(self, results: List[dict], existing_results: List[dict] = None):
        self.ocr_button.setEnabled(True); self.ocr_button.setText("🔍 开始OCR识别")
        self.progress_bar.setVisible(False)
        if self.is_selecting_mask: self.toggle_mask_selection(False)
        
        # 合并相邻的OCR结果，如+0.03这种被分成多个部分的情况
        print(f"🔍 开始合并OCR结果，原始结果数量: {len(results)}")
        results = self.merge_adjacent_ocr_results(results)
        print(f"✅ 合并完成，合并后结果数量: {len(results)}")
        
        # 合并现有的OCR结果和新的结果
        if existing_results:
            # 创建一个集合来检查重复项
            existing_boxes = set()
            for r in existing_results:
                if 'bbox' in r:
                    # 将bbox转换为可哈希格式以检查重复
                    bbox_tuple = tuple(tuple(point) for point in r['bbox'])
                    existing_boxes.add(bbox_tuple)
            
            # 过滤掉重叠的新结果
            new_results = []
            for r in results:
                if 'bbox' in r:
                    bbox_tuple = tuple(tuple(point) for point in r['bbox'])
                    if bbox_tuple not in existing_boxes:
                        # 检测是否为竖排文本 - 检查边界框的高度和宽度
                        if 'is_vertical' not in r and 'bbox' in r and len(r['bbox']) >= 4:
                            points = np.array(r['bbox'])
                            min_x, min_y = np.min(points, axis=0)
                            max_x, max_y = np.max(points, axis=0)
                            width = max_x - min_x
                            height = max_y - min_y
                            
                            # 如果高度大于等于宽度，标记为竖排文本
                            if height >= width:
                                r['is_vertical'] = True
                                print(f"✅ 自动检测到竖排文本 (尺寸: {width:.1f}x{height:.1f}, 高宽比: {height/width:.2f})")
                        
                        new_results.append(r)
                else:
                    new_results.append(r)
            
            # 合并结果
            self.ocr_results = existing_results + new_results
        else:
            # 为新结果检测竖排文本
            for r in results:
                if 'is_vertical' not in r and 'bbox' in r and len(r['bbox']) >= 4:
                    points = np.array(r['bbox'])
                    min_x, min_y = np.min(points, axis=0)
                    max_x, max_y = np.max(points, axis=0)
                    width = max_x - min_x
                    height = max_y - min_y
                    
                    # 如果高度大于等于宽度，标记为竖排文本
                    if height >= width:
                        r['is_vertical'] = True
                        print(f"✅ 自动检测到竖排文本 (尺寸: {width:.1f}x{height:.1f}, 高宽比: {height/width:.2f})")
            
            self.ocr_results = results
        
        # 如果是多页PDF，保存当前页的OCR结果
        if self.pdf_file_path and self.current_pdf_page in range(self.pdf_page_count):
            self.ocr_results_by_page[self.current_pdf_page] = self.ocr_results.copy()
            
        self.update_ocr_stats()
        self.display_ocr_results()
        
        # 显示带有时间信息的完成消息
        message = f"成功识别出 {len(results)} 个文本区域。\n您可以选择创建标注或进一步筛选结果。"
        
        # 添加文件名
        if self.current_file_path:
            message += f"\n\n文件: {Path(self.current_file_path).name}"
        
        # 添加日期时间
        now = datetime.now()
        message += f"\n完成时间: {now.strftime('%Y-%m-%d %H:%M:%S')}"
        
        QMessageBox.information(self, "OCR识别完成", message)

    def update_ocr_stats(self):
        total_count = len(self.ocr_results)
        type_counts = {}
        for result in self.ocr_results: 
            result_type = result.get('type', 'unknown')
            type_counts[result_type] = type_counts.get(result_type, 0) + 1
        stats_text = f"识别结果: {total_count}个文本"
        if type_counts: 
            stats_text += f" ({', '.join([f'{k}({v})' for k, v in type_counts.items()])})"
        self.ocr_stats_label.setText(stats_text)

    def display_ocr_results(self):
        self.clear_ocr_display()
        for i, result in enumerate(self.ocr_results): 
            self.create_ocr_bbox_item(result, i)

    def create_ocr_bbox_item(self, ocr_result, index):
        if not HAS_OCR_SUPPORT: return
        bbox = ocr_result['bbox']
        
        bbox_array = np.array(bbox)
        path = QPainterPath()
        path.moveTo(bbox_array[0][0], bbox_array[0][1])
        for point in bbox_array[1:]: 
            path.lineTo(point[0], point[1])
        path.closeSubpath()
        
        # 使用可调整大小的路径项
        from ui.graphics_view import ResizableGraphicsPathItem
        bbox_item = ResizableGraphicsPathItem(path)
        
        # 统一使用淡蓝色，不再区分类型或竖排文本
        color = QColor(91, 192, 235, 120)  # 淡蓝色，与annotation类型相同
        bbox_item.setPen(QPen(color, 2))
        bbox_item.setBrush(QBrush(color))
        
        # 设置自定义属性以便识别
        bbox_item.setData(Qt.UserRole, 10000 + index)  # 使用10000+索引作为标识
        bbox_item.setData(Qt.UserRole + 1, ocr_result)  # 存储OCR结果
        
        # 存储原始边界框和OCR结果到自定义类
        bbox_item.original_bbox = bbox
        bbox_item.ocr_result = ocr_result
        bbox_item.associated_annotations = []  # 初始化关联的气泡标注列表
        
        # 连接信号
        bbox_item.bbox_updated.connect(self.on_bbox_updated)
        
        self.graphics_scene.addItem(bbox_item)
        return bbox_item

    def clear_ocr_display(self):
        """清除OCR结果的可视化显示"""
        try:
            # 找出所有OCR边界框项目并移除
            bbox_items = []
            for item in self.graphics_scene.items():
                # 检查是否为OCR边界框类型的项目
                if item.data(Qt.UserRole) is not None and isinstance(item.data(Qt.UserRole), int) and item.data(Qt.UserRole) >= 10000:
                    bbox_items.append(item)
            
            # 从场景中删除找到的边界框项目
            for item in bbox_items:
                try:
                    self.graphics_scene.removeItem(item)
                except Exception as e:
                    print(f"移除OCR边界框时出错: {e}")
        except Exception as e:
            print(f"清除OCR显示时出错: {e}")

    def clear_ocr_results(self):
        """清除OCR结果"""
        try:
            self.clear_ocr_display()
            self.ocr_results = []
            
            # 如果是多页PDF，清除当前页的OCR结果缓存
            if self.pdf_file_path and self.current_pdf_page in range(self.pdf_page_count):
                if self.current_pdf_page in self.ocr_results_by_page:
                    self.ocr_results_by_page[self.current_pdf_page] = []
                    
            self.update_ocr_stats()
        except Exception as e:
            print(f"清除OCR结果时出错: {e}")
            # 确保OCR结果被清空
            self.ocr_results = []

    def filter_ocr_results(self):
        """筛选OCR结果，仅显示符合条件的结果"""
        try:
            # 获取筛选条件
            filter_type = self.filter_combo.currentText()
            
            # 根据筛选条件获取结果
            if filter_type == "全部": 
                filtered_results = self.ocr_results
            else:
                target_type = OCR_FILTER_TYPE_MAP.get(filter_type, "annotation")
                filtered_results = [r for r in self.ocr_results if r.get('type', 'annotation') == target_type]
            
            # 清除当前OCR显示
            self.clear_ocr_display()
            
            # 显示过滤后的结果
            for i, result in enumerate(filtered_results):
                if result in self.ocr_results:
                    original_index = self.ocr_results.index(result)
                    self.create_ocr_bbox_item(result, original_index)
                    
            # 更新状态
            self.status_bar.showMessage(f"筛选后显示 {len(filtered_results)}/{len(self.ocr_results)} 个OCR结果", 3000)
        except Exception as e:
            print(f"筛选OCR结果时出错: {e}")
            # 出错时显示全部结果
            self.display_ocr_results()

    def create_annotations_from_ocr(self):
        """从OCR结果创建标注"""
        if not self.ocr_results:
            QMessageBox.information(self, "提示", "请先进行OCR识别")
            return
            
        # 获取设置的置信度阈值
        confidence_threshold = self.confidence_slider.value() / 100.0
        
        # 统计创建了多少个新标注
        created_count = 0
        for result in self.ocr_results:
            # 只处理置信度高于阈值的结果
            if result.get('confidence', 0) >= confidence_threshold:
                annotation = self.create_annotation_from_ocr_result(result)
                if annotation:
                    created_count += 1
                
        if created_count > 0:
            # 如果是多页PDF模式，更新当前页面的标注缓存
            if self.pdf_file_path and self.current_pdf_page in range(self.pdf_page_count):
                # 使用新的保存方法更新当前页的标注数据
                self.save_current_page_data()
                
            QMessageBox.information(self, "完成", f"已创建 {created_count} 个标注")
        else:
            QMessageBox.information(self, "提示", "没有符合置信度要求的OCR结果可创建标注")
            
        # 刷新标注列表
        self.refresh_annotation_list()

    def _parse_annotation_text(self, text: str) -> dict:
        """解析标注文本，提取尺寸、类型和公差信息"""
        result = {
            'dimension': '',
            'dimension_type': '',
            'upper_tolerance': '',
            'lower_tolerance': ''
        }
        
        # 处理空文本
        if not text:
            return result
            
        # 清理文本，移除括号内容
        text_main = re.sub(r'\s*\(.*\)', '', text).strip()
        
        print(f"🔍 开始解析文本: '{text_main}'")
        
        # 特殊处理合并后的"Φ7 0 +0.02"格式
        # 先检查是否是这种特殊格式
        special_match = re.match(r'([Φ∅øMR])\s*(\d+\.?\d*)\s+0\s+([+\-][\d\.]+)', text_main, re.IGNORECASE)
        if special_match:
            symbol = special_match.group(1).upper()
            if symbol in ['Φ', '∅', 'Ø']:
                result['dimension_type'] = 'Φ'
            elif symbol == 'R':
                result['dimension_type'] = 'R'
            elif symbol == 'M':
                result['dimension_type'] = 'M'
                
            result['dimension'] = special_match.group(2)
            result['lower_tolerance'] = '0'
            
            # 处理公差值
            tolerance = special_match.group(3)
            if tolerance.startswith('+'):
                result['upper_tolerance'] = tolerance
            elif tolerance.startswith('-'):
                result['lower_tolerance'] = tolerance
                
            print(f"  ✓ 特殊格式匹配成功: 符号='{symbol}', 尺寸='{result['dimension']}', 上公差='{result['upper_tolerance']}', 下公差='{result['lower_tolerance']}'")
            return result
            
        # 检查直径符号
        match = re.match(r'([Φ∅øMR])\s*(\d+\.?\d*)(.*)', text_main, re.IGNORECASE)
        if match:
            symbol = match.group(1).upper()
            if symbol in ['Φ', '∅', 'Ø']:
                result['dimension_type'] = 'Φ'
            elif symbol == 'R':
                result['dimension_type'] = 'R'
            elif symbol == 'M':
                result['dimension_type'] = 'M'
                
            result['dimension'] = match.group(2)
            remaining_text = match.group(3).strip()
            print(f"  ✓ 识别到符号: '{symbol}', 尺寸: '{result['dimension']}', 剩余文本: '{remaining_text}'")
        else:
            # 没有特殊符号，提取数字作为尺寸
            number_match = re.search(r'(\d+\.?\d*)', text_main)
            if number_match:
                result['dimension'] = number_match.group(1)
                remaining_text = text_main[number_match.end():].strip()
                print(f"  ✓ 识别到尺寸: '{result['dimension']}', 剩余文本: '{remaining_text}'")
                
                # 检查是否包含角度符号，自动设置为角度类型
                if '°' in remaining_text or '度' in remaining_text:
                    result['dimension_type'] = '∠'  # 设置为角度类型符号
                    print(f"  ✓ 检测到角度符号，自动设置尺寸类型为: '∠'")
            else:
                remaining_text = text_main
                print(f"  ⚠ 未识别到尺寸, 剩余文本: '{remaining_text}'")
        
        # 处理合并后的公差格式 (例如: "0 +0.02" 或 "+0.02 0" 或 "0 +0.02 -0.01")
        # 先查找是否有独立的0作为基准公差
        zero_match = re.search(r'\b0\b', remaining_text)
        plus_match = re.search(r'\+(\d+\.?\d*)', remaining_text)
        minus_matches = re.findall(r'\-(\d+\.?\d*)', remaining_text)
        
        # 处理±格式的公差 (例如: 83.02±0.01)
        if '±' in remaining_text:
            parts = remaining_text.split('±')
            if len(parts) == 2:
                tolerance_value = parts[1].strip()
                if tolerance_value:
                    result['upper_tolerance'] = '+' + tolerance_value
                    result['lower_tolerance'] = '-' + tolerance_value
                    print(f"  ✓ 识别到±公差: '{tolerance_value}', 上公差: '{result['upper_tolerance']}', 下公差: '{result['lower_tolerance']}'")
        
        # 处理上下公差格式
        else:
            # 提取上公差
            if plus_match:
                result['upper_tolerance'] = '+' + plus_match.group(1)
                print(f"  ✓ 识别到上公差: '{result['upper_tolerance']}'")
            
            # 处理有多个负号公差的情况（例如：两个-号表示的公差值）
            if len(minus_matches) >= 2:
                # 将负号后面的数值转为浮点数进行比较
                minus_values = [float(val) for val in minus_matches]
                
                # 找出较大值（数值绝对值较小）和较小值（数值绝对值较大）
                min_value = min(minus_values)
                max_value = max(minus_values)
                
                # 将较小的值作为上公差（负值中，绝对值大的数值更小）
                result['upper_tolerance'] = f"-{min_value}"
                # 将较大的值作为下公差（负值中，绝对值小的数值更大）
                result['lower_tolerance'] = f"-{max_value}"
                
                print(f"  ✓ 识别到多个负号公差: '{minus_matches}', 通过比较大小确定：上公差='{result['upper_tolerance']}', 下公差='{result['lower_tolerance']}'")
                
            # 处理单个负号公差
            elif len(minus_matches) == 1:
                result['lower_tolerance'] = '-' + minus_matches[0]
                print(f"  ✓ 识别到下公差: '{result['lower_tolerance']}'")
            
            # 如果有0且没有负公差，将0设为下公差
            elif zero_match and not result['lower_tolerance']:
                result['lower_tolerance'] = '0'
                print(f"  ✓ 识别到下公差为0")
        
        # 最终结果输出
        print(f"📋 解析结果: 尺寸: '{result['dimension']}', 类型: '{result['dimension_type']}', 上公差: '{result['upper_tolerance']}', 下公差: '{result['lower_tolerance']}'")
        
        return result
        
    def create_annotation_from_ocr_result(self, ocr_result: dict):
        """从OCR结果创建标注"""
        if 'bbox' not in ocr_result or len(ocr_result['bbox']) < 4:
            return None
            
        # 调试输出 - 检查OCR结果
        print(f"\n🔍 调试OCR结果:")
        for key, value in ocr_result.items():
            if key == 'bbox':
                print(f"  - {key}: [包含{len(value)}个点的边界框]")
            else:
                print(f"  - {key}: {value}")
        
        # 提取文本和置信度
        text = ocr_result.get('text', '')
        confidence = ocr_result.get('confidence', 0.0)
        
        # 检查是否为竖排文本
        is_vertical = ocr_result.get('is_vertical', False)
        
        # 预处理文本 - 处理常见的OCR错误
        # 1. 将"O"替换为"0"，如果它看起来像是数字
        text = re.sub(r'(?<!\w)O(?!\w)', '0', text)
        
        # 2. 规范化直径符号
        text = text.replace('Ø', 'Φ').replace('∅', 'Φ')
        
        print(f"📝 预处理后文本: '{text}'")
        
        # 获取边界框中心点
        center_x = ocr_result.get('center_x')
        center_y = ocr_result.get('center_y')
        
        if center_x is None or center_y is None:
            # 如果没有中心点，则从bbox计算
            bbox = ocr_result['bbox']
            x_coords = [point[0] for point in bbox]
            y_coords = [point[1] for point in bbox]
            center_x = sum(x_coords) / len(x_coords)
            center_y = sum(y_coords) / len(y_coords)
            
        # 尝试分析文本内容
        parsed_data = self._parse_annotation_text(text)
        
        # 创建标注
        anchor_point = QPointF(center_x, center_y)
        
        # 如果结果中没有type字段，确保添加一个默认值
        if 'type' not in ocr_result:
            # 尝试根据文本内容分类
            from core.paddle_ocr_worker import PaddleOCRWorker
            ocr_result['type'] = PaddleOCRWorker._classify_mechanical_text(None, text)
        
        # 检查type字段并使用
        text_type = ocr_result.get('type', 'annotation')
        
        # 统一使用淡蓝色样式，不再根据类型区分颜色
        style = 'default'
        
        print(f"📌 创建标注: 文本='{text}', 类型='{text_type}', 样式='{style}'")
        
        # 确保parsed_data中包含从OCR结果获取的维度类型信息
        if not parsed_data.get('dimension_type'):
            if text_type == 'diameter' or 'Φ' in text:
                parsed_data['dimension_type'] = 'Φ'
                print(f"  ✓ 根据类型设置尺寸类型为: 'Φ'")
            elif text_type == 'thread_spec' or text.startswith('M'):
                parsed_data['dimension_type'] = 'M'
                print(f"  ✓ 根据类型设置尺寸类型为: 'M'")
            elif text_type == 'angle' or '°' in text:
                parsed_data['dimension_type'] = '∠'
                print(f"  ✓ 根据类型设置尺寸类型为: '∠'")
            else:
                # 默认设置为直线度
                parsed_data['dimension_type'] = '⏤'
                print(f"  ✓ 设置默认尺寸类型为: '⏤'")
        
        # 如果没有解析出尺寸，但文本中有数字，尝试提取
        if not parsed_data.get('dimension'):
            # 提取第一个数字序列作为尺寸
            dimension_match = re.search(r'\d+\.?\d*', text)
            if dimension_match:
                parsed_data['dimension'] = dimension_match.group(0)
                print(f"  ✓ 从文本中提取尺寸: '{parsed_data['dimension']}'")
        
        # 创建标注对象
        annotation = self._create_new_annotation(
            anchor_point, 
            text,
            parsed_data.get('dimension', ''), 
            parsed_data.get('dimension_type', ''),
            style
        )
        
        # 设置公差值
        if parsed_data.get('upper_tolerance'):
            annotation.set_upper_tolerance(parsed_data.get('upper_tolerance'))
            print(f"  ✓ 设置上公差: '{parsed_data.get('upper_tolerance')}'")
        if parsed_data.get('lower_tolerance'):
            annotation.set_lower_tolerance(parsed_data.get('lower_tolerance'))
            print(f"  ✓ 设置下公差: '{parsed_data.get('lower_tolerance')}'")
        
        # 设置边界框点信息，以便箭头能够指向文本框边缘
        if annotation and 'bbox' in ocr_result:
            bbox_points = []
            for point in ocr_result['bbox']:
                bbox_points.append(QPointF(point[0], point[1]))
            annotation.set_bbox_points(bbox_points)
            
            # 找到对应的OCR框图形项并建立关联
            for item in self.graphics_scene.items():
                if hasattr(item, 'ocr_result') and item.ocr_result is ocr_result:
                    # 找到了对应的OCR框
                    if hasattr(item, 'associated_annotations'):
                        # 将当前气泡标注添加到OCR框的关联列表
                        item.associated_annotations.append(annotation)
                    break
        
        print(f"✅ 标注创建完成: ID={annotation.annotation_id}, 尺寸={annotation.dimension}, 类型={annotation.dimension_type}, 上公差={annotation.upper_tolerance}, 下公差={annotation.lower_tolerance}")
        
        return annotation

    def _create_new_annotation(self, anchor_point: QPointF, text: str = "", dimension: str = "", dimension_type: str = "", style: str = "default"):
        # 在创建新标注前，先确定全局最大ID
        if self.pdf_file_path and self.pdf_page_count > 1:
            # 多页PDF模式，计算所有页面中的最大ID
            max_id_across_pages = self.annotation_counter
            
            # 遍历所有页面的标注数据
            for page_idx in range(self.pdf_page_count):
                if page_idx in self.annotations_by_page and self.annotations_by_page[page_idx]:
                    page_annotations = self.annotations_by_page[page_idx]
                    if page_annotations:
                        page_max_id = max(annotation_data['annotation_id'] for annotation_data in page_annotations)
                        max_id_across_pages = max(max_id_across_pages, page_max_id)
            
            # 更新计数器为全局最大值
            self.annotation_counter = max_id_across_pages
            
        # 递增标注计数器
        self.annotation_counter += 1
        
        shape_map = {"空心圆": "circle", "实心圆": "solid_circle", "五角星": "pentagram", "三角形": "triangle"}
        selected_shape = shape_map.get(self.shape_combo.currentText(), "circle")
        
        # 创建标注项（初始大小为15，后续会调整）
        annotation = BubbleAnnotationItem(
            annotation_id=self.annotation_counter,
            anchor_point=anchor_point,
            text=text,
            style=style,
            shape=selected_shape,
            color=self.next_annotation_color,
            size=15,  # 临时大小，会基于scale_factor调整
            dimension=dimension,
            dimension_type=dimension_type
        )
        
        # 设置比例因子并触发大小自动计算
        annotation.auto_radius = True
        annotation.scale_factor = self.next_annotation_scale
        
        # 使用-1触发自动计算大小
        annotation.change_size(-1)
        
        # 清除颜色设置（一次性的）
        if self.next_annotation_color: self.next_annotation_color = None
        
        annotation.data_updated.connect(self.annotation_table.update_annotation_data)
        annotation.selected.connect(self.on_annotation_selected)
        annotation.moved.connect(self.on_annotation_moved)
        annotation.delete_requested.connect(self.delete_annotation)
        annotation.style_change_requested.connect(self.on_annotation_style_changed)
        annotation.shape_change_requested.connect(self.on_annotation_shape_changed)
        annotation.color_change_requested.connect(self.on_annotation_color_changed)
        annotation.size_change_requested.connect(self.on_annotation_size_changed)
        
        self.graphics_scene.addItem(annotation)
        self.annotations.append(annotation)
        self.annotation_table.add_annotation(annotation, {})
        
        return annotation

    def on_annotation_selected(self, annotation: BubbleAnnotationItem):
        for ann in self.annotations: ann.set_highlighted(False)
        self.current_annotation = annotation; annotation.set_highlighted(True)
        
        # 获取真实图像位置预览区域
        preview_rect = self.get_annotation_preview_rect(annotation)
        self.property_editor.set_annotation(annotation, self.current_pixmap, preview_rect)
        
        self.annotation_table.highlight_annotation(annotation.annotation_id)
        style_text = "自定义" if annotation.custom_color else STYLE_NAME_MAP.get(annotation.style, "默认")
        self.style_combo.blockSignals(True); self.style_combo.setCurrentText(style_text); self.style_combo.blockSignals(False)
        shape_map_rev = {"circle": "空心圆", "solid_circle": "实心圆", "pentagram": "五角星", "triangle": "三角形"}
        shape_text = shape_map_rev.get(annotation.shape_type, "空心圆")
        self.shape_combo.blockSignals(True); self.shape_combo.setCurrentText(shape_text); self.shape_combo.blockSignals(False)
        
        # 更新滑块和输入框为当前比例
        self.size_slider.blockSignals(True)
        self.size_input.blockSignals(True)
        scale_percent = int(annotation.scale_factor * 100)
        self.size_slider.setValue(scale_percent)
        self.size_input.setText(str(scale_percent))
        self.size_label.setText(f"{scale_percent}%")
        self.size_slider.blockSignals(False)
        self.size_input.blockSignals(False)
        
        self.update_color_button_display()
        
    def get_annotation_preview_rect(self, annotation: BubbleAnnotationItem):
        """获取标注预览区域的矩形"""
        # 尝试查找关联的OCR结果
        if hasattr(annotation, 'bbox_points') and annotation.bbox_points:
            # 如果标注有bbox信息，直接使用
            bbox_points = annotation.bbox_points
            x_values = [p.x() for p in bbox_points]
            y_values = [p.y() for p in bbox_points]
            min_x = min(x_values)
            min_y = min(y_values)
            max_x = max(x_values)
            max_y = max(y_values)
            width = max_x - min_x
            height = max_y - min_y
            
            # 稍微扩大一点区域，方便查看
            padding = max(width, height) * 0.2
            preview_rect = QRectF(
                min_x - padding,
                min_y - padding,
                width + padding * 2,
                height + padding * 2
            )
            print(f"从bbox获取预览区域: {preview_rect}")
            return preview_rect
        
        # 如果没有bbox，尝试根据OCR结果查找
        ocr_results = self._find_matching_ocr_results(annotation.anchor_point, annotation.text)
        if ocr_results and len(ocr_results) > 0:
            # 使用OCR边界框
            best_match = ocr_results[0]
            if 'bbox' in best_match and best_match['bbox']:
                bbox = best_match['bbox']
                # 计算边界框的边界
                x_values = [point[0] for point in bbox]
                y_values = [point[1] for point in bbox]
                min_x = min(x_values)
                min_y = min(y_values)
                max_x = max(x_values)
                max_y = max(y_values)
                width = max_x - min_x
                height = max_y - min_y
                
                # 稍微扩大一点区域，方便查看
                padding = max(width, height) * 0.2
                preview_rect = QRectF(
                    min_x - padding,
                    min_y - padding,
                    width + padding * 2,
                    height + padding * 2
                )
                print(f"从OCR结果获取预览区域: {preview_rect}")
                return preview_rect
        
        # 如果没有关联OCR结果，使用锚点为中心的默认区域
        anchor_pos = annotation.anchor_point
        default_size = 100  # 默认区域大小
        preview_rect = QRectF(
            anchor_pos.x() - default_size / 2,
            anchor_pos.y() - default_size / 2,
            default_size,
            default_size
        )
        print(f"使用默认区域: {preview_rect}")
        return preview_rect
    
    def on_annotation_moved(self, annotation: BubbleAnnotationItem, position: QPointF):
        if annotation == self.current_annotation:
            # 获取新的预览区域并更新
            preview_rect = self.get_annotation_preview_rect(annotation)
            self.property_editor.preview_rect = preview_rect
            self.property_editor.update_preview()
    
    def select_annotation_by_id(self, annotation_id: int):
        """根据ID选中标注
        
        Args:
            annotation_id: 要选中的标注ID
        """
        try:
            # 首先确保annotations列表有效
            if not self.annotations:
                return
                
            # 查找对应ID的标注
            found_annotation = None
            for annotation in self.annotations:
                if annotation.annotation_id == annotation_id:
                    found_annotation = annotation
                    break
                    
            # 如果找不到对应ID的标注，直接返回
            if not found_annotation:
                return
                
            # 添加安全检查，确保对象仍然有效
            try:
                # 尝试访问对象的一个属性，如果对象已删除会抛出异常
                _ = found_annotation.isVisible()
            except RuntimeError:
                print(f"警告: 标注 #{annotation_id} 对象已被删除，无法选中")
                return
                
            # 将视图中心对准标注
            self.graphics_view.centerOn(found_annotation)
            self.graphics_scene.clearSelection()
            found_annotation.setSelected(True)
            self.on_annotation_selected(found_annotation)
        except Exception as e:
            print(f"选中标注时出错: {e}")
            # 不向用户显示错误，静默失败
    
    def toggle_area_selection(self, checked: bool):
        self.graphics_view.set_selection_mode(checked)
        self.area_select_action.setText("退出区域OCR" if checked else "区域OCR标注")
        
        # 禁用或启用其他可能冲突的功能
        if checked:
            # 如果启用区域选择，禁用屏蔽区域选择
            if self.mask_select_action.isChecked():
                self.mask_select_action.blockSignals(True)
                self.mask_select_action.setChecked(False)
                self.mask_select_action.blockSignals(False)
                self.is_selecting_mask = False
            
            # 显示提示信息
            self.status_bar.showMessage("区域OCR识别模式：请在图纸上拖拽选择要识别的区域...", 5000)
        else:
            self.status_bar.showMessage("已退出区域OCR识别模式", 3000)
    
    def create_annotation_in_area(self, rect: QRectF):
        if self.is_selecting_mask: return
        self._create_new_annotation(
            anchor_point=rect.center(),
            text=f"区域标注 {self.annotation_counter + 1}"
        )
        self.area_select_action.setChecked(False)
    
    def delete_annotation(self, annotation: BubbleAnnotationItem):
        """删除标注，并同步删除关联的OCR结果和数据模型"""
        try:
            annotation_id_to_delete = annotation.annotation_id

            # 检查此标注是否有关联的OCR边界框
            if hasattr(annotation, 'bbox_points') and annotation.bbox_points:
                # 将QPointF列表转换为可比较的元组列表
                bbox_to_find = [
                    (round(p.x(), 2), round(p.y(), 2)) for p in annotation.bbox_points
                ]
                bbox_to_find.sort()

                # 查找并删除匹配的OCR结果
                ocr_result_to_remove = None
                for ocr_result in self.ocr_results:
                    if 'bbox' in ocr_result:
                        # 将OCR结果的bbox也转换为标准格式进行比较
                        current_bbox = [
                            (round(p[0], 2), round(p[1], 2)) for p in ocr_result['bbox']
                        ]
                        current_bbox.sort()
                        
                        if bbox_to_find == current_bbox:
                            ocr_result_to_remove = ocr_result
                            break
                
                if ocr_result_to_remove:
                    self.ocr_results.remove(ocr_result_to_remove)
                    # 重新显示OCR结果以移除高亮框
                    self.clear_ocr_display()
                    self.display_ocr_results()
                    self.update_ocr_stats()
                    print(f"成功删除与标注 #{annotation_id_to_delete} 关联的OCR结果。")

            # 从场景和当前活动列表中删除标注对象
            self.graphics_scene.removeItem(annotation)
            self.annotations.remove(annotation)
            
            # --- 关键修复 ---
            # 直接从数据模型 (self.annotations_by_page) 中移除该标注的数据
            if self.pdf_file_path and self.current_pdf_page in self.annotations_by_page:
                page_data = self.annotations_by_page[self.current_pdf_page]
                # 寻找并删除具有相同ID的标注数据
                self.annotations_by_page[self.current_pdf_page] = [
                    ann_data for ann_data in page_data if ann_data.get('annotation_id') != annotation_id_to_delete
                ]
                print(f"已从第 {self.current_pdf_page + 1} 页的数据模型中删除标注 #{annotation_id_to_delete}")

            if self.current_annotation == annotation:
                self.current_annotation = None
                self.property_editor.set_annotation(None, None, None)
            
            # 更新标注列表
            self.refresh_annotation_list()
            
            self.status_bar.showMessage(f"已删除标注 #{annotation_id_to_delete}", 2000)
        except Exception as e:
            print(f"删除标注时出错: {e}")
            self.status_bar.showMessage(f"删除标注失败: {str(e)}", 2000)
    
    def refresh_annotation_list(self):
        # 使用新的排序方法，直接将所有标注传递给表格进行排序和显示
        self.annotation_table.sort_annotations(self.annotations)
    
    def on_annotation_style_changed(self, annotation: BubbleAnnotationItem):
        if annotation == self.current_annotation: self.on_annotation_selected(annotation)

    def change_current_annotation_style(self, style_text: str):
        if self.current_annotation and style_text != "自定义":
            new_style = STYLE_NAME_REVERSE_MAP.get(style_text, "default")
            self.current_annotation.change_style(new_style)

    def on_annotation_shape_changed(self, annotation: BubbleAnnotationItem):
        if annotation == self.current_annotation: self.on_annotation_selected(annotation)

    def change_current_annotation_shape(self, shape_text: str):
        if self.current_annotation:
            shape_map = {"空心圆": "circle", "实心圆": "solid_circle", "五角星": "pentagram", "三角形": "triangle"}
            new_shape = shape_map.get(shape_text, "circle")
            self.current_annotation.change_shape(new_shape)

    def on_annotation_color_changed(self, annotation: BubbleAnnotationItem):
        if annotation == self.current_annotation: self.on_annotation_selected(annotation)
        
    def reorder_annotations(self):
        """重新排序所有气泡标注
        
        按照从左到右，从上到下的顺序重新对所有气泡标注进行编号
        在多页PDF模式下，考虑前面页面的标注数量，保持连续编号
        """
        if not self.annotations:
            QMessageBox.information(self, "提示", "没有标注可以重新排序")
            return
            
        # 确认对话框
        confirm_message = "确定要重新排序当前页面的气泡标注吗？"
        
        # 判断是否为多页PDF模式
        if self.pdf_file_path and self.pdf_page_count > 1:
            confirm_message = "检测到多页PDF，请选择重新排序方式：\n\n" \
                             "【是】: 从页面1开始全局重排序（跨页面重新从1开始编号）\n" \
                             "【否】: 仅重排序当前页面（考虑前面页面的标注数量）\n" \
                             "【取消】: 取消操作"
                             
            confirm = QMessageBox.question(
                self, 
                "确认重新排序", 
                confirm_message,
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.No
            )
            
            if confirm == QMessageBox.Cancel:
                return
                
            if confirm == QMessageBox.Yes:
                # 全局重排序 - 从页面1开始重新编号
                return self._reorder_all_pdf_pages()
                
            # 否则继续当前页面排序，但考虑前面页面的标注数量
            return self._reorder_current_page_with_continuity()
        else:
            # 非PDF模式，或单页PDF
            confirm = QMessageBox.question(
                self, 
                "确认重新排序", 
                confirm_message,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if confirm != QMessageBox.Yes:
                return
                
            # 普通排序
            self._reorder_current_page(start_id=1)
            
    def _reorder_current_page(self, start_id=1):
        """重排序当前页面的标注
        
        Args:
            start_id: 起始ID
        """
        # 改进的排序方法：使用精确的Y坐标而非网格，确保更准确的从上到下排序
        sorted_annotations = sorted(
            self.annotations,
            key=lambda ann: (ann.scenePos().y(), ann.scenePos().x())
        )
        
        # 保存当前选中的标注
        current_annotation_id = self.current_annotation.annotation_id if self.current_annotation else None
        
        # 重新分配ID
        new_id = start_id
        for annotation in sorted_annotations:
            old_id = annotation.annotation_id
            annotation.annotation_id = new_id
            
            # 更新文本（如果文本中包含ID）
            if str(old_id) in annotation.text:
                annotation.text = annotation.text.replace(str(old_id), str(new_id))
                
            # 更新气泡显示
            annotation.update_annotation_id_display()
            
            # 发送数据更新信号
            annotation.data_updated.emit(annotation)
            
            new_id += 1
        
        # 更新标注计数器
        # 之前的逻辑 `self.annotation_counter = max(self.annotation_counter, new_id - 1)` 是错误的，
        # 因为它不允许计数器在删除项目后减小。
        # 正确的做法是，在重新编号后，将计数器设置为新的最大ID。
        # 对于单页排序，这等于 new_id - 1。
        # 对于多页连续排序，这也正确地设定了当前所有已知项中的最大ID。
        self.annotation_counter = new_id - 1
        
        # 刷新标注列表
        self.refresh_annotation_list()
        
        # 更新当前页面的缓存
        if self.pdf_file_path and self.current_pdf_page in range(self.pdf_page_count):
            self.save_current_page_data()
        
        # 恢复选中的标注（如果可能）
        if current_annotation_id is not None:
            # 尝试找到原来的标注
            for annotation in self.annotations:
                if annotation.annotation_id == current_annotation_id:
                    self.select_annotation_by_id(annotation.annotation_id)
                    break
        
        self.status_bar.showMessage(f"已成功重新排序 {len(sorted_annotations)} 个气泡标注（从上到下，从左到右）", 3000)
            
    def _reorder_current_page_with_continuity(self):
        """重新排序当前页面，保持与前面页面的连续性"""
        # 计算前面页面的标注数量总和
        previous_annotations_count = 0
        
        if self.pdf_file_path and self.current_pdf_page > 0:
            for page_idx in range(self.current_pdf_page):
                if page_idx in self.annotations_by_page:
                    previous_annotations_count += len(self.annotations_by_page[page_idx])
        
        # 从前面页面标注数量+1开始编号
        start_id = previous_annotations_count + 1
        
        # 执行排序
        self._reorder_current_page(start_id=start_id)
        
        self.status_bar.showMessage(f"已重新排序当前页面标注，起始编号: {start_id}，保持与前面页面的连续性", 3000)
        
    def _reorder_all_pdf_pages(self):
        """全局重排序所有PDF页面【重构版】"""
        if not self.pdf_file_path:
            return

        # 1. 保存当前页的任何未保存的更改
        self.save_current_page_data()
        current_page_before_reorder = self.current_pdf_page

        try:
            # 2. 从数据模型中收集所有页面的所有标注数据
            all_annotations_data = []
            page_map = {} # 用于将标注数据映射回其原始页面
            
            for page_idx in range(self.pdf_page_count):
                if page_idx in self.annotations_by_page:
                    for ann_data in self.annotations_by_page[page_idx]:
                        # 添加页面索引，以便后续识别
                        ann_data['_page_idx'] = page_idx
                        all_annotations_data.append(ann_data)

            if not all_annotations_data:
                QMessageBox.information(self, "提示", "所有页面都没有标注可重新排序。")
                return

            # 3. 对所有标注数据进行排序
            # 排序规则：首先按页面索引，然后按Y坐标，最后按X坐标
            sorted_data = sorted(
                all_annotations_data,
                key=lambda data: (data.get('_page_idx', 0), data.get('pos_y', 0), data.get('pos_x', 0))
            )

            # 4. 重新分配ID
            next_id = 1
            for ann_data in sorted_data:
                old_id = ann_data['annotation_id']
                ann_data['annotation_id'] = next_id
                
                # 如果旧ID存在于文本中，也一并更新
                if str(old_id) in ann_data['text']:
                    ann_data['text'] = ann_data['text'].replace(str(old_id), str(next_id))
                
                next_id += 1

            # 5. 将更新后的数据重新组织回按页码的字典结构中
            self.annotations_by_page.clear()
            for ann_data in sorted_data:
                page_idx = ann_data.pop('_page_idx') # 移除临时页面索引
                if page_idx not in self.annotations_by_page:
                    self.annotations_by_page[page_idx] = []
                self.annotations_by_page[page_idx].append(ann_data)

            # 6. 更新全局标注计数器
            self.annotation_counter = next_id - 1

            # 7. 重新加载用户之前所在的页面以显示更新
            # load_pdf_page会处理场景清理和从新数据重建标注
            self.load_pdf_page(current_page_before_reorder, skip_save=True)

            QMessageBox.information(self, "全局重排序完成", f"已完成所有 {self.pdf_page_count} 页PDF的标注重排序，总标注数量: {next_id-1}")

        except Exception as e:
            logger.exception("全局重排序过程中发生严重错误")
            QMessageBox.warning(self, "重排序出错", f"全局重排序过程中发生意外错误: {str(e)}\n\n建议重启程序。")
            # 尝试恢复到原始页面
            self.load_pdf_page(current_page_before_reorder, skip_save=True)

    def select_annotation_color(self):
        initial_color = QColor("blue")
        if self.current_annotation and self.current_annotation.custom_color:
            initial_color = self.current_annotation.custom_color
        elif self.next_annotation_color:
            initial_color = self.next_annotation_color
        color = QColorDialog.getColor(initial_color, self, "选择标注颜色")
        if color.isValid():
            if self.current_annotation:
                self.current_annotation.change_color(color)
            else:
                self.next_annotation_color = color
                self.update_color_button_display()
                self.status_bar.showMessage(f"下一个标注的颜色已设置为 {color.name()}", 3000)

    def update_color_button_display(self):
        color_to_show = None
        if self.current_annotation and self.current_annotation.custom_color:
            color_to_show = self.current_annotation.custom_color
        elif self.next_annotation_color:
            color_to_show = self.next_annotation_color
        if color_to_show and color_to_show.isValid():
            self.color_button.setStyleSheet(f"QPushButton {{ background-color: {color_to_show.name()}; color: {'white' if color_to_show.lightnessF() < 0.5 else 'black'}; border: 1px solid grey; font-weight: bold; }}")
        else:
            self.color_button.setStyleSheet("")
            self.next_annotation_color = None

    def on_annotation_size_changed(self, annotation: BubbleAnnotationItem):
        """当标注的大小发生变化时的回调
        
        当某个气泡的大小变化时，直接更新UI，但不触发更多气泡大小变化
        """
        # 获取当前气泡的比例因子
        scale_factor = annotation.scale_factor
        percent = int(scale_factor * 100)
        
        # 仅更新UI显示，不触发其他气泡的更新
        self.size_slider.blockSignals(True)
        self.size_slider.setValue(percent)
        self.size_slider.blockSignals(False)
        
        # 同步更新输入框
        self.size_input.blockSignals(True)
        self.size_input.setText(str(percent))
        self.size_input.blockSignals(False)
        
        # 更新标签文本
        self.size_label.setText(f"{percent}%")
        
        # 保存为全局设置，但不应用到其他气泡
        self.next_annotation_scale = scale_factor
        self.next_annotation_size = -1
        
        # 仅刷新属性编辑器
        if annotation == self.current_annotation:
            self.property_editor.update_preview()

    def change_annotation_size(self, percent: int):
        """更改所有注释的大小（基于全局统一设置）
        
        Args:
            percent: 大小比例，范围50-160（对应50%-160%）
        """
        # 调试信息，帮助追踪滑块问题
        print(f"滑块值变化: {percent}%")
        
        # 显示百分比文本
        self.size_label.setText(f"{percent}%")
        
        # 同步更新输入框（如果调用来源是滑块）
        if self.sender() == self.size_slider:
            self.size_input.blockSignals(True)
            self.size_input.setText(str(percent))
            self.size_input.blockSignals(False)
        
        # 将百分比转换为比例因子，比如100%=1.0，50%=0.5
        scale_factor = percent / 100.0
        
        # 保存为全局比例设置
        self.next_annotation_scale = scale_factor
        self.next_annotation_size = -1  # 标记为自动
        
        # 直接修改场景中的所有气泡
        updated_count = 0
        
        # 禁用所有更新和重绘
        self.graphics_scene.blockSignals(True)
        
        # 批量更新所有气泡的参数，但不触发信号和更新
        for annotation in self.annotations:
            try:
                # 直接设置参数，不调用任何可能递归的方法
                annotation.blockSignals(True)  # 阻止信号传播
                annotation.scale_factor = scale_factor
                annotation.auto_radius = True
                annotation.base_radius = 20  # 统一使用固定基准半径
                annotation.radius = max(int(annotation.base_radius * scale_factor), 10)
                updated_count += 1
            except Exception as e:
                print(f"设置气泡参数时出错: {e}")
        
        # 批量处理完成后，解除信号阻塞并触发一次场景更新
        for annotation in self.annotations:
            try:
                annotation.blockSignals(False)
                annotation.prepareGeometryChange()
            except Exception as e:
                print(f"解除信号阻塞时出错: {e}")
        
        # 解除场景信号阻塞
        self.graphics_scene.blockSignals(False)
        
        # 强制更新场景
        self.graphics_scene.update()
        
        # 如果当前有选中的气泡，刷新属性编辑器
        if self.current_annotation:
            self.property_editor.update_preview()
            
        self.status_bar.showMessage(f"已更新 {updated_count} 个气泡的大小为 {percent}%", 3000)

    def clear_annotations(self, show_empty_message=True):
        """清除所有标注"""
        if not self.annotations:
            if show_empty_message:
                self.status_bar.showMessage("没有标注可清除", 2000)
            return

        confirm = QMessageBox.question(
            self, "确认清除", "确定要删除所有标注吗？这个操作不能撤销。", 
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) if show_empty_message else QMessageBox.Yes

        if confirm == QMessageBox.Yes:
            # 先清空标注表，避免引用已删除对象
            self.annotation_table.clear_annotations()
            
            # 清除当前选中状态
            self.current_annotation = None
            self.property_editor.set_annotation(None, None, None)
            
            # 创建一个临时副本，避免在迭代时修改列表
            annotations_to_remove = list(self.annotations)
            
            # 清空标注列表先，这样我们在删除场景项时就不会尝试访问self.annotations中的对象了
            self.annotations.clear()
            
            # 从场景中删除标注
            for annotation in annotations_to_remove:
                try:
                    self.graphics_scene.removeItem(annotation)
                except Exception as e:
                    print(f"删除标注时出错: {e}")
                    # 继续处理其他标注
            
            # 清空副本以释放引用
            annotations_to_remove.clear()
            
            # 当前是否为多页PDF模式
            if self.pdf_file_path and self.current_pdf_page in range(self.pdf_page_count):
                # 清除当前页面的缓存标注
                if self.current_pdf_page in self.annotations_by_page:
                    self.annotations_by_page[self.current_pdf_page] = []
            
            if show_empty_message:
                self.status_bar.showMessage("已清除所有标注", 2000)

    def toggle_mask_selection(self, checked: bool):
        self.is_selecting_mask = checked; self.graphics_view.set_selection_mode(checked)
        if hasattr(self, 'mask_select_action'):
            self.mask_select_action.blockSignals(True); self.mask_select_action.setChecked(checked); self.mask_select_action.blockSignals(False)
        if checked and hasattr(self, 'area_select_action'):
            self.area_select_action.blockSignals(True); self.area_select_action.setChecked(False); self.area_select_action.blockSignals(False)
        self.status_bar.showMessage("屏蔽区域选择模式：拖拽鼠标选择要屏蔽的区域" if checked else "已退出屏蔽区域选择模式", 3000 if not checked else 0)
    
    def handle_area_selection(self, rect: QRectF):
        if self.is_selecting_mask:
            self.add_masked_region(rect)
        else:
            # 修改为对选中区域进行OCR识别
            self.run_ocr_on_selected_area(rect)
    
    def add_masked_region(self, rect: QRectF):
        self.masked_regions.append(rect)
        self.display_masked_region(rect, len(self.masked_regions) - 1)
        self.update_mask_count()
        self.status_bar.showMessage(f"已添加屏蔽区域 {len(self.masked_regions)}", 2000)
    
    def display_masked_region(self, rect: QRectF, index: int):
        from PySide6.QtWidgets import QGraphicsRectItem
        mask_item = QGraphicsRectItem(rect)
        mask_color = QColor(255, 0, 0, 80); border_color = QColor(255, 0, 0, 200)
        mask_item.setPen(QPen(border_color, 2, Qt.DashLine)); mask_item.setBrush(QBrush(mask_color))
        mask_item.mask_region_index = index; mask_item.setZValue(100)
        self.graphics_scene.addItem(mask_item)
    
    def clear_masked_regions(self):
        self.masked_regions.clear()
        items_to_remove = [item for item in self.graphics_scene.items() if hasattr(item, 'mask_region_index')]
        for item in items_to_remove:
            self.graphics_scene.removeItem(item)
        self.update_mask_count()
        self.status_bar.showMessage("已清除所有屏蔽区域", 2000)
    
    def update_mask_count(self): pass
    
    def is_point_in_masked_region(self, x: float, y: float) -> bool:
        return any(region.contains(QPointF(x, y)) for region in self.masked_regions)
    
    def is_bbox_in_masked_region(self, bbox) -> bool:
        if not self.masked_regions: return False
        if hasattr(bbox, '__len__') and len(bbox) >= 4:
            if HAS_OCR_SUPPORT:
                x_min, y_min = np.min(bbox, axis=0)
                x_max, y_max = np.max(bbox, axis=0)
            else:
                x_coords, y_coords = [p[0] for p in bbox], [p[1] for p in bbox]
                x_min, x_max, y_min, y_max = min(x_coords), max(x_coords), min(y_coords), max(y_coords)
            bbox_rect = QRectF(x_min, y_min, x_max - x_min, y_max - y_min)
        else:
            bbox_rect = bbox
        return any(region.intersects(bbox_rect) for region in self.masked_regions)

    def on_gpu_checkbox_toggled(self, checked):
        """当GPU复选框状态变化时，更新CPU复选框状态"""
        if checked and self.cpu_checkbox.isChecked():
            self.cpu_checkbox.blockSignals(True)
            self.cpu_checkbox.setChecked(False)
            self.cpu_checkbox.blockSignals(False)
            # 禁用线程数输入框
            self.threads_spinbox.setEnabled(False)

    def on_cpu_checkbox_toggled(self, checked):
        """当CPU复选框状态变化时，更新GPU复选框状态"""
        if checked and self.gpu_checkbox.isChecked():
            self.gpu_checkbox.blockSignals(True)
            self.gpu_checkbox.setChecked(False)
            self.gpu_checkbox.blockSignals(False)
        # 根据CPU选择状态启用/禁用线程数输入框    
        self.threads_spinbox.setEnabled(checked)

    def change_current_annotation_text(self, new_text: str):
        """修改当前选中标注的文本内容"""
        if self.current_annotation and self.current_annotation.text != new_text:
            self.current_annotation.set_text(new_text)
            return True
        return False

    def on_bbox_updated(self, bbox_item):
        """处理OCR框更新事件，更新关联的气泡标注"""
        try:
            if not hasattr(bbox_item, 'associated_annotations') or not bbox_item.associated_annotations:
                return
                
            # 获取更新后的边界框
            if not hasattr(bbox_item, 'ocr_result') or 'bbox' not in bbox_item.ocr_result:
                return
            
            # 获取当前时间戳，用于防抖保护
            current_time = time.time()
            last_update = getattr(self, '_last_annotation_update_time', 0)
            
            # 检查是否是调整大小引起的更新
            is_resize_update = hasattr(bbox_item, '_update_from_resize') and bbox_item._update_from_resize
            
            # 非调整大小更新时应用防抖
            if not is_resize_update and (current_time - last_update) * 1000 < 50:  # 50毫秒内不重复处理
                return
            
            self._last_annotation_update_time = current_time
            
            # 调整大小后重置标志
            if is_resize_update:
                bbox_item._update_from_resize = False
                
            # 更新所有关联的气泡标注
            for annotation in bbox_item.associated_annotations:
                try:
                    # 创建QPointF列表
                    bbox_points = []
                    for point in bbox_item.ocr_result['bbox']:
                        if isinstance(point, (list, tuple)) and len(point) >= 2:
                            bbox_points.append(QPointF(point[0], point[1]))
                    
                    if len(bbox_points) >= 4:
                        # 设置新的边界框点
                        annotation.set_bbox_points(bbox_points)
                        
                        # 确保锚点也一起更新（在边界框中心）
                        # 计算新的边界框中心点
                        x_sum = sum(p.x() for p in bbox_points)
                        y_sum = sum(p.y() for p in bbox_points)
                        center_x = x_sum / len(bbox_points)
                        center_y = y_sum / len(bbox_points)
                        
                        # 更新锚点位置
                        annotation.anchor_point = QPointF(center_x, center_y)
                        
                        # 强制更新气泡几何形状
                        annotation.prepareGeometryChange()
                        annotation._update_geometry()
                        annotation.update()
                except Exception as e:
                    print(f"更新气泡标注 {annotation.annotation_id} 时出错: {e}")
        except Exception as e:
            print(f"处理OCR框更新事件时出错: {e}")

    def run_ocr_on_selected_area(self, rect: QRectF):
        """对选中区域进行OCR识别"""
        if not HAS_OCR_SUPPORT:
            QMessageBox.warning(self, "功能缺失", "OCR功能需要PaddleOCR和依赖包。请安装所需依赖。")
            self.create_annotation_in_area(rect)  # 仍然创建区域标注
            return
        
        if not self.current_pixmap:
            QMessageBox.information(self, "提示", "请先打开图片文件。")
            return
        
        # 始终使用直接识别模式，不调用检测模型
        direct_recognition = True
            
        # 从当前图像中截取选定区域
        x, y, width, height = rect.x(), rect.y(), rect.width(), rect.height()
        
        # 确保坐标在有效范围内
        x = max(0, int(x))
        y = max(0, int(y))
        width = min(int(width), self.current_pixmap.width() - x)
        height = min(int(height), self.current_pixmap.height() - y)
        
        # 创建临时文件保存选定区域
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            temp_path = temp_file.name
        
        # 截取并保存区域图像
        cropped_pixmap = self.current_pixmap.copy(x, y, width, height)
        
        # 自动判断是否为竖排文本 - 如果高度大于等于宽度，则认为是竖排文本
        is_vertical = height >= width  # 高度大于等于宽度时认为是竖排文本
        
        if is_vertical:
            print(f"✅ 自动检测到竖排文本 (尺寸: {width}x{height}, 高宽比: {height/width:.2f})")
            self.status_bar.showMessage(f"检测到竖排文本，将自动旋转识别 (高宽比: {height/width:.2f})", 3000)
        else:
            print(f"✅ 自动检测到横排文本 (尺寸: {width}x{height}, 高宽比: {height/width:.2f})")
            self.status_bar.showMessage(f"检测到横排文本 (高宽比: {height/width:.2f})", 3000)
        
        # 保存处理后的图像用于OCR
        cropped_pixmap.save(temp_path)
        
        self.status_bar.showMessage("正在对选中区域进行直接OCR识别（跳过检测模型）...")
        
        # 获取语言配置
        lang_text = self.language_combo.currentText()
        lang_code = DEFAULT_OCR_LANGUAGES.get(lang_text, ["ch_sim"])
        
        # 获取环境配置
        force_cpu = self.cpu_checkbox.isChecked()
        use_gpu = self.gpu_checkbox.isChecked() and not force_cpu
        
        # 获取CPU线程数
        cpu_threads = self.threads_spinbox.value()
        
        # 创建区域OCR工作器
        self.area_ocr_worker = PaddleOCRWorker(
            temp_path, 
            lang_code, 
            [],  # 区域识别不需要屏蔽区域
            force_cpu=force_cpu,
            cpu_threads=cpu_threads,  # 传递线程数
            direct_recognition=direct_recognition  # 设置为直接识别模式，跳过检测模型
        )
        
        # 设置竖排文本标记
        if is_vertical:
            self.area_ocr_worker.is_vertical_text = True
        
        # 保存原始选择区域信息，供结果处理使用
        self.area_ocr_worker.original_rect = {
            'x': x,
            'y': y,
            'width': width,
            'height': height
        }
        
        # 连接信号
        self.area_ocr_worker.signals.progress.connect(lambda p: self.progress_bar.setValue(p))
        self.area_ocr_worker.signals.error.connect(self.on_area_ocr_error)
        
        # 使用lambda捕获rect参数，传递给回调函数
        # 传递额外的参数is_vertical，以便在结果处理时考虑旋转
        self.area_ocr_worker.signals.finished.connect(
            lambda results: self.on_area_ocr_finished(results, rect, temp_path, x, y, is_vertical)
        )
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 启动线程
        self.thread_pool.start(self.area_ocr_worker)
        
    def on_area_ocr_error(self, error_msg: str):
        """区域OCR错误处理"""
        self.progress_bar.setVisible(False)
        QMessageBox.warning(self, "区域OCR识别错误", error_msg)
        self.area_select_action.setChecked(False)
    
    def on_area_ocr_finished(self, results: List[dict], rect: QRectF, temp_path: str, offset_x: int, offset_y: int, is_vertical: bool):
        """区域OCR完成处理"""
        self.progress_bar.setVisible(False)
        
        # 删除临时文件
        try:
            os.unlink(temp_path)
        except:
            pass
        
        # 如果没有识别结果，创建空白标注
        if not results:
            QMessageBox.information(self, "区域OCR", "选中区域未识别到文字，将创建空白标注。")
            self.create_annotation_in_area(rect)
            self.area_select_action.setChecked(False)
            return
            
        # 合并相邻的OCR结果，如+0.03这种被分成多个部分的情况
        results = self.merge_adjacent_ocr_results(results)
        
        # 调整结果坐标（添加偏移量）
        for result in results:
            if 'bbox' in result:
                adjusted_bbox = []
                
                # 获取选择区域的宽度和高度，用于竖排文本的坐标转换
                rect_width = int(rect.width())
                rect_height = int(rect.height())
                
                for point in result['bbox']:
                    if is_vertical:
                        # 竖排文本情况下，需要将旋转后的坐标转换回原始坐标系统
                        # 图像旋转了90度顺时针，经过多次尝试修正后的变换公式：
                        original_x = offset_x + point[1]
                        original_y = offset_y + point[0]  # 简化变换，尝试直接使用旋转后的x作为y的偏移
                        adjusted_bbox.append([original_x, original_y])
                    else:
                        # 正常情况下只添加偏移量
                        adjusted_bbox.append([point[0] + offset_x, point[1] + offset_y])
                
                result['bbox'] = adjusted_bbox
            
            if 'center_x' in result and 'center_y' in result:
                if is_vertical:
                    # 同样需要转换中心点坐标
                    original_center_x = offset_x + result['center_y']
                    original_center_y = offset_y + result['center_x']
                    result['center_x'] = original_center_x
                    result['center_y'] = original_center_y
                else:
                    result['center_x'] += offset_x
                    result['center_y'] += offset_y
        
        # 创建底色显示区域 - 与全局OCR一样显示识别区域
        for i, result in enumerate(results):
            self.create_ocr_bbox_item(result, i)
                
        # 为每个OCR结果创建标注
        confidence_threshold = self.confidence_slider.value() / 100.0
        created_count = 0
        
        for result in results:
            if result.get('confidence', 0) >= confidence_threshold:
                # 修改为使用相对于场景的正确坐标创建标注
                annotation = self.create_annotation_from_ocr_result(result)
                if annotation:
                    created_count += 1
        
        # 将识别结果添加到全局OCR结果中，以便筛选和管理
        self.ocr_results.extend(results)
        self.update_ocr_stats()
        
        if created_count > 0:
            QMessageBox.information(self, "区域OCR完成", f"在选中区域内识别出 {len(results)} 个文本，创建了 {created_count} 个标注。")
            self.refresh_annotation_list()
        else:
            QMessageBox.information(self, "区域OCR", "选中区域的识别结果未达到置信度阈值，将创建空白标注。")
            self.create_annotation_in_area(rect)
        
        self.area_select_action.setChecked(False)

    def delete_current_annotation(self):
        """删除当前选中的标注"""
        if not self.current_annotation:
            QMessageBox.information(self, "提示", "请先选择一个标注")
            return
            
        confirm = QMessageBox.question(
            self, "确认删除", f"确定要删除标注 #{self.current_annotation.annotation_id} 吗？", 
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if confirm == QMessageBox.Yes:
            # 保存当前标注引用，以便删除
            annotation_to_delete = self.current_annotation
            self.current_annotation = None
            self.property_editor.set_annotation(None, None, None)
            
            # 删除标注及其关联的OCR结果
            self.delete_annotation(annotation_to_delete)

    def _find_matching_ocr_results(self, anchor_point, annotation_text):
        """多策略匹配OCR结果 - 优化版，更严格的标准避免误匹配"""
        matching_results = []
        
        # 从文本中提取原始OCR文本（如果有）
        original_ocr_text = None
        if annotation_text and "原始文本:" in annotation_text:
            parts = annotation_text.split("原始文本:")
            if len(parts) > 1:
                original_ocr_text = parts[1].strip()
        
        # 确保只找到最匹配的一个OCR结果
        best_match = None
        best_match_score = float('inf')  # 分数越小越匹配
        
        # 遍历所有OCR结果
        for i, ocr_result in enumerate(self.ocr_results):
            current_score = float('inf')  # 初始化为最大值
            
            # 策略1: 精确文本匹配 - 如果标注中包含原始OCR文本，则检查是否完全匹配
            if original_ocr_text and 'text' in ocr_result:
                ocr_text = ocr_result['text']
                if ocr_text == original_ocr_text:
                    # 完全匹配，这是最优先级
                    matching_results = [ocr_result]
                    return matching_results
                elif ocr_text.strip() == original_ocr_text.strip():
                    # 除了空格外完全匹配
                    matching_results = [ocr_result]
                    return matching_results
            
            # 策略2: 位置匹配 - 当没有完全文本匹配时，计算最近的一个
            if 'bbox' in ocr_result:
                bbox = ocr_result['bbox']
                if len(bbox) >= 4:
                    # 计算OCR框的中心点和边界
                    x_coords = [point[0] for point in bbox]
                    y_coords = [point[1] for point in bbox]
                    center_x = sum(x_coords) / len(bbox)
                    center_y = sum(y_coords) / len(bbox)
                    ocr_center = QPointF(center_x, center_y)
                    
                    # 计算OCR中心到标注锚点的距离作为分数
                    distance = ((ocr_center.x() - anchor_point.x())**2 + 
                                (ocr_center.y() - anchor_point.y())**2)**0.5
                    
                    # 也考虑特殊的右侧位置关系（加权）
                    right_edge = max(x_coords)
                    if anchor_point.x() > right_edge and abs(anchor_point.y() - center_y) < 20:
                        # 如果位置关系很明确（标注在OCR右侧），距离分数减半
                        distance *= 0.5
                    
                    # 更新得分
                    current_score = distance
            
            # 如果这个OCR结果比之前找到的更匹配，更新最佳匹配
            if current_score < best_match_score:
                best_match_score = current_score
                best_match = ocr_result
        
        # 只有当最佳匹配的距离小于阈值时才返回结果
        # 使用固定阈值80像素，更严格的匹配标准
        if best_match is not None and best_match_score < 80:
            matching_results.append(best_match)
        
        return matching_results

    def convert_pdf_to_images(self):
        """将PDF文件批量转换为PNG图片"""
        if not HAS_OCR_SUPPORT:
            QMessageBox.warning(self, "功能缺失", "PDF转换功能需要PyMuPDF支持。请安装所需依赖。")
            return
            
        # 打开文件选择对话框，仅选择PDF文件
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("PDF文件 (*.pdf)")
        if not file_dialog.exec():
            return
            
        pdf_paths = file_dialog.selectedFiles()
        if not pdf_paths:
            return
            
        pdf_path = pdf_paths[0]
        pdf_filename = Path(pdf_path).name
        
        # 获取质量设置
        zoom_factor = PDF_QUALITY_OPTIONS.get(self.pdf_quality_combo.currentText(), 4.0)
        
        # 显示正在处理的消息
        self.status_bar.showMessage(f"正在处理PDF: {pdf_filename}...")
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        try:
            # 调用转换函数
            png_paths, error = FileLoader.convert_pdf_to_pngs(pdf_path, zoom_factor=zoom_factor)
            
            # 隐藏进度条
            self.progress_bar.setVisible(False)
            
            if error:
                QMessageBox.warning(self, "转换失败", f"PDF转换失败: {error}")
                self.status_bar.showMessage(f"❌ PDF转换失败: {error}", 5000)
                return
                
            if not png_paths:
                QMessageBox.warning(self, "转换失败", "未能生成PNG文件。")
                self.status_bar.showMessage("❌ PDF转换失败: 未能生成PNG文件", 5000)
                return
                
            # 转换成功，显示成功消息
            success_message = f"PDF成功转换为{len(png_paths)}个PNG文件：\n\n"
            for i, path in enumerate(png_paths[:5]):  # 只显示前5个文件路径
                success_message += f"{i+1}. {path}\n"
                
            if len(png_paths) > 5:
                success_message += f"\n... 以及另外 {len(png_paths) - 5} 个文件"
                
            # 询问是否打开第一个生成的PNG文件
            result = QMessageBox.information(
                self, 
                "转换成功", 
                success_message + "\n\n是否打开第一个PNG文件？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if result == QMessageBox.Yes and png_paths:
                # 加载第一个PNG文件
                self.load_file(png_paths[0])
            else:
                self.status_bar.showMessage(f"✅ PDF转换完成: {len(png_paths)}个文件", 5000)
                
        except Exception as e:
            self.progress_bar.setVisible(False)
            QMessageBox.critical(self, "错误", f"转换过程中发生错误: {str(e)}")
            self.status_bar.showMessage(f"❌ PDF转换错误: {str(e)}", 5000)

    def load_pdf_page(self, page_index: int, skip_save: bool = False):
        """加载指定页码的PDF页面
        
        Args:
            page_index: 页码（从0开始）
            skip_save: 是否跳过保存当前页数据的步骤。用于在执行了外部数据修改（如全局重排）后刷新视图。
            
        Returns:
            bool: 是否成功启动加载过程
        """
        if not self.pdf_file_path or page_index not in range(self.pdf_page_count):
            return False
        
        start_time = time.time()
        logger.debug(f"开始加载PDF页面 {page_index+1}/{self.pdf_page_count}")
        
        # 设置加载对话框文本并显示
        self.loading_label.setText(f"⏳ 正在加载第 {page_index+1}/{self.pdf_page_count} 页...\n请稍候")
        self.loading_dialog.resize(self.size())
        # 非阻塞方式显示
        self.loading_dialog.show()
        QApplication.processEvents()  # 确保UI立即更新
        logger.debug(f"显示加载对话框耗时: {time.time() - start_time:.2f}秒")
        
        # 记录之前的页码，以便加载失败时可以恢复
        self.previous_page = self.current_pdf_page
        
        # 清除当前选择状态，防止引用已删除的对象
        self.graphics_scene.clearSelection()
        self.current_annotation = None
        self.property_editor.set_annotation(None, None, None)
        logger.debug(f"清除当前选择状态耗时: {time.time() - start_time:.2f}秒")
        
        if not skip_save:
            # 保存当前页面的标注和OCR结果
            if self.current_pdf_page in range(self.pdf_page_count):
                save_start = time.time()
                self.save_current_page_data()
                logger.debug(f"保存当前页面数据耗时: {time.time() - save_start:.2f}秒")
        else:
            logger.debug("跳过页面数据保存（按需刷新模式）")
        
        # 更新当前页码
        self.current_pdf_page = page_index
        
        # 更新导航按钮状态
        self.update_pdf_navigation_controls()
        
        # 检查是否已经缓存了该页面
        if page_index in self.pdf_pages_cache:
            temp_path = self.pdf_pages_cache[page_index]
            # 检查临时文件是否仍然存在
            if Path(temp_path).exists():
                logger.debug(f"发现页面缓存: {temp_path}")
                self.status_bar.showMessage(f"从缓存加载PDF页面 {page_index+1}/{self.pdf_page_count}...")
                self.loading_label.setText(f"正在从缓存加载第 {page_index+1}/{self.pdf_page_count} 页...\n请稍候")
                QApplication.processEvents()  # 确保UI立即更新
                
                # 在后台线程中加载缓存图像
                def load_cached_image():
                    try:
                        cache_start = time.time()
                        logger.debug(f"开始从缓存加载图像...")
                        pixmap = QPixmap(temp_path)
                        if not pixmap.isNull():
                            logger.debug(f"缓存图像加载成功，耗时: {time.time() - cache_start:.2f}秒")
                            # 在主线程中更新UI
                            QApplication.instance().postEvent(self, LoadPDFEvent(pixmap, temp_path))
                            return True
                        else:
                            logger.error(f"缓存图像加载失败: pixmap为空")
                    except Exception as e:
                        logger.exception(f"加载缓存图像出错: {e}")
                        return False
                
                # 创建线程并启动
                thread = threading.Thread(target=load_cached_image)
                thread.daemon = True
                logger.debug(f"启动缓存图像加载线程")
                thread.start()
                return True
        
        # 缓存中没有或临时文件已被删除，重新转换
        zoom_factor = PDF_QUALITY_OPTIONS.get(self.pdf_quality_combo.currentText(), 4.0)
        logger.debug(f"未找到缓存，开始转换PDF，缩放因子: {zoom_factor}")
        self.status_bar.showMessage(f"正在转换PDF页面 {page_index+1}/{self.pdf_page_count}...")
        
        # 更新加载提示
        self.loading_label.setText(f"正在转换第 {page_index+1}/{self.pdf_page_count} 页...\n请稍候")
        QApplication.processEvents()  # 确保UI立即更新
        
        # 清除当前场景
        self.graphics_scene.clear()
        
        # 创建PDF加载工作线程
        pdf_loader = PDFLoaderWorker(
            self.pdf_file_path,
            page_index,
            quality=zoom_factor,
            force_resolution=self.force_resolution_checkbox.isChecked()
        )
        
        # 连接信号
        pdf_loader.signals.finished.connect(self._on_pdf_loaded)
        pdf_loader.signals.error.connect(self._on_pdf_load_error)
        
        # 启动线程
        logger.debug(f"启动PDF加载线程，总准备耗时: {time.time() - start_time:.2f}秒")
        self.thread_pool.start(pdf_loader)
        
        return True

    def _on_pdf_loaded(self, pixmap: QPixmap, temp_path: str):
        """PDF加载完成处理"""
        try:
            start_time = time.time()
            logger.debug(f"PDF加载完成回调开始处理")
            
            # 缓存此页面
            self.pdf_pages_cache[self.current_pdf_page] = temp_path
            self.current_pixmap = pixmap
            self.current_file_path = temp_path
            
            # 清除并设置场景
            self.graphics_scene.clear()
            self.graphics_scene.addPixmap(pixmap)
            logger.debug(f"更新场景耗时: {time.time() - start_time:.2f}秒")
            
            # 强制处理事件，确保场景已更新
            QApplication.processEvents()
            
            # 仅在初始加载时执行一次居中操作
            self.center_view()
            
            # 恢复此页面的数据
            restore_start = time.time()
            self.restore_page_data(self.current_pdf_page)
            logger.debug(f"恢复页面数据耗时: {time.time() - restore_start:.2f}秒")
            
            self.status_bar.showMessage(f"✅ 页面加载成功: {self.current_pdf_page+1}/{self.pdf_page_count}", 3000)
            
            # 隐藏加载对话框
            self.loading_dialog.hide()
            logger.debug(f"PDF加载完成处理总耗时: {time.time() - start_time:.2f}秒")
        except Exception as e:
            logger.exception(f"PDF加载完成处理中发生异常: {str(e)}")
            self._on_pdf_load_error(f"处理加载结果时出错: {str(e)}")

    def _on_pdf_load_error(self, error_msg: str):
        """PDF加载出错处理"""
        logger.error(f"PDF加载错误: {error_msg}")
        # 恢复到之前的页面
        self.current_pdf_page = self.previous_page
        self.update_pdf_navigation_controls()
        
        # 如果当前场景是空的，尝试恢复之前的页面内容
        if len(self.graphics_scene.items()) == 0 and self.previous_page in self.pdf_pages_cache:
            try:
                logger.debug(f"尝试恢复到之前的页面: {self.previous_page+1}")
                prev_temp_path = self.pdf_pages_cache[self.previous_page]
                if Path(prev_temp_path).exists():
                    try:
                        prev_pixmap = QPixmap(prev_temp_path)
                        self.graphics_scene.addPixmap(prev_pixmap)
                        self.current_pixmap = prev_pixmap
                        self.current_file_path = prev_temp_path
                        
                        # 直接居中显示，不使用延迟
                        self.graphics_view.centerContent()
                        
                        self.restore_page_data(self.previous_page)
                        logger.debug(f"成功恢复到之前的页面")
                    except Exception as e:
                        logger.exception(f"恢复之前页面时出错: {str(e)}")
                        pass  # 如果恢复失败，至少保持当前状态
            except Exception as e:
                logger.exception(f"尝试恢复之前页面时出错: {str(e)}")
        
        QMessageBox.warning(self, "错误", f"加载PDF页面失败: {error_msg}")
        self.status_bar.showMessage(f"❌ 页面加载失败: {error_msg}", 3000)
        
        # 隐藏加载对话框
        self.loading_dialog.hide()

    def update_pdf_navigation_controls(self):
        """更新PDF导航控件的状态"""
        if not self.pdf_file_path:
            self.pdf_nav_widget.setVisible(False)
            return
            
        self.pdf_nav_widget.setVisible(self.pdf_page_count > 1)
        
        # 更新页码显示
        self.page_label.setText(f"{self.current_pdf_page+1}/{self.pdf_page_count}")
        
        # 更新导航按钮状态
        self.prev_page_btn.setEnabled(self.current_pdf_page > 0)
        self.next_page_btn.setEnabled(self.current_pdf_page < self.pdf_page_count - 1)
        self.go_to_page_btn.setEnabled(self.pdf_page_count > 1)
            
    def save_current_page_data(self):
        """保存当前页面的标注和OCR结果"""
        if self.current_pdf_page not in range(self.pdf_page_count):
            return
            
        # 保存当前页的OCR结果
        try:
            self.ocr_results_by_page[self.current_pdf_page] = self.ocr_results.copy()
        except Exception as e:
            print(f"保存OCR结果时出错: {e}")
            # 确保有一个空列表
            self.ocr_results_by_page[self.current_pdf_page] = []
        
        # 保存当前页的标注数据（不是对象引用）
        annotation_data_list = []
        
        # 安全地获取标注数据
        for annotation in list(self.annotations):  # 使用列表副本进行迭代
            try:
                # 检查对象是否有效
                if not annotation.scene():
                    print(f"警告: 标注 #{getattr(annotation, 'annotation_id', 'unknown')} 不在场景中，跳过保存")
                    continue
                
                # 提取标注的基本属性
                annotation_data = {
                    'annotation_id': annotation.annotation_id,
                    'text': annotation.text,
                    'style': annotation.style,
                    'shape_type': annotation.shape_type,
                    'radius': annotation.radius,
                    'base_radius': annotation.base_radius,
                    'scale_factor': annotation.scale_factor,
                    'dimension': annotation.dimension,
                    'dimension_type': annotation.dimension_type,
                    'upper_tolerance': annotation.upper_tolerance,
                    'lower_tolerance': annotation.lower_tolerance,
                    'is_audited': annotation.is_audited,
                    'auto_radius': annotation.auto_radius,
                    'pos_x': annotation.pos().x(),
                    'pos_y': annotation.pos().y(),
                    'anchor_x': annotation.anchor_point.x(),
                    'anchor_y': annotation.anchor_point.y(),
                }
                
                # 保存颜色信息
                if annotation.custom_color and annotation.custom_color.isValid():
                    annotation_data['color'] = {
                        'r': annotation.custom_color.red(),
                        'g': annotation.custom_color.green(),
                        'b': annotation.custom_color.blue(),
                        'a': annotation.custom_color.alpha(),
                    }
                else:
                    annotation_data['color'] = None
                    
                # 保存边界框点信息
                if hasattr(annotation, 'bbox_points') and annotation.bbox_points:
                    bbox_points_data = []
                    for point in annotation.bbox_points:
                        bbox_points_data.append((point.x(), point.y()))
                    annotation_data['bbox_points'] = bbox_points_data
                else:
                    annotation_data['bbox_points'] = []
                    
                annotation_data_list.append(annotation_data)
            except Exception as e:
                print(f"保存标注数据时出错: {e}")
                # 继续处理下一个标注
            
        # 存储数据字典而不是对象引用
        self.annotations_by_page[self.current_pdf_page] = annotation_data_list

    def restore_page_data(self, page_index: int):
        """恢复指定页面的标注和OCR结果"""
        # 先清空标注表，避免引用已删除对象
        self.annotation_table.clear_annotations()
        
        # 清除当前选中状态
        self.current_annotation = None
        self.property_editor.set_annotation(None, None, None)
        
        # 清理场景中的所有标注对象和OCR边界框
        items_to_remove = []
        for item in self.graphics_scene.items():
            # 删除标注对象和OCR边界框
            if isinstance(item, BubbleAnnotationItem) or \
               (item.data(Qt.UserRole) is not None and isinstance(item.data(Qt.UserRole), int) and item.data(Qt.UserRole) >= 10000):
                items_to_remove.append(item)
        
        # 从场景中移除所有标注和OCR边界框
        for item in items_to_remove:
            try:
                self.graphics_scene.removeItem(item)
            except Exception as e:
                print(f"移除项目时出错: {e}")
                # 继续处理
        
        # 清空当前标注列表和OCR结果
        self.annotations = []
        self.ocr_results = []
        
        # 恢复标注
        if page_index in self.annotations_by_page and self.annotations_by_page[page_index]:
            annotation_data_list = self.annotations_by_page[page_index]
            
            # 根据保存的数据重新创建标注对象
            for annotation_data in annotation_data_list:
                try:
                    # 创建位置
                    position = QPointF(annotation_data['pos_x'], annotation_data['pos_y'])
                    anchor_point = QPointF(annotation_data['anchor_x'], annotation_data['anchor_y'])
                    
                    # 创建颜色对象
                    color = None
                    if annotation_data['color']:
                        color_data = annotation_data['color']
                        color = QColor(
                            color_data['r'],
                            color_data['g'],
                            color_data['b'],
                            color_data['a']
                        )
                    
                    # 创建新的标注对象
                    annotation = BubbleAnnotationItem(
                        annotation_id=annotation_data['annotation_id'],
                        anchor_point=anchor_point,
                        text=annotation_data['text'],
                        style=annotation_data['style'],
                        shape=annotation_data['shape_type'],
                        color=color,
                        size=annotation_data['radius'],
                        dimension=annotation_data['dimension'],
                        dimension_type=annotation_data['dimension_type'],
                        upper_tolerance=annotation_data['upper_tolerance'],
                        lower_tolerance=annotation_data['lower_tolerance'],
                        is_audited=annotation_data['is_audited']
                    )
                    
                    # 设置其他属性
                    annotation.setPos(position)
                    annotation.base_radius = annotation_data['base_radius']
                    annotation.scale_factor = annotation_data['scale_factor']
                    annotation.auto_radius = annotation_data['auto_radius']
                    
                    # 恢复边界框点
                    if annotation_data['bbox_points']:
                        bbox_points = []
                        for point_tuple in annotation_data['bbox_points']:
                            bbox_points.append(QPointF(point_tuple[0], point_tuple[1]))
                        annotation.set_bbox_points(bbox_points)
                    
                    # 连接信号
                    self._connect_annotation_signals(annotation)
                    
                    # 添加到场景和列表
                    self.graphics_scene.addItem(annotation)
                    self.annotations.append(annotation)
                except Exception as e:
                    print(f"恢复标注时出错: {e}")
                    # 继续处理下一个标注
                
            # 更新标注计数器
            if self.annotations:
                self.annotation_counter = max(annotation.annotation_id for annotation in self.annotations)
                
        # 恢复OCR结果
        if page_index in self.ocr_results_by_page and self.ocr_results_by_page[page_index]:
            self.ocr_results = self.ocr_results_by_page[page_index].copy()
            
            # 重新显示OCR边界框
            self.display_ocr_results()
            self.update_ocr_stats()
        
        # 最后再刷新标注列表，确保使用的是当前页面的标注
        self.refresh_annotation_list()

    def _connect_annotation_signals(self, annotation):
        """连接标注对象的所有信号"""
        annotation.selected.connect(self.on_annotation_selected)
        annotation.moved.connect(self.on_annotation_moved)
        annotation.delete_requested.connect(self.delete_annotation)
        annotation.size_change_requested.connect(self.on_annotation_size_changed)
        annotation.shape_change_requested.connect(self.on_annotation_shape_changed)
        annotation.style_change_requested.connect(self.on_annotation_style_changed)
        annotation.color_change_requested.connect(self.on_annotation_color_changed)
        annotation.data_updated.connect(lambda: self.refresh_annotation_list())

    def go_to_prev_page(self):
        """转到上一页"""
        if self.pdf_file_path and self.current_pdf_page > 0:
            self.load_pdf_page(self.current_pdf_page - 1)
            
    def go_to_next_page(self):
        """转到下一页"""
        if self.pdf_file_path and self.current_pdf_page < self.pdf_page_count - 1:
            self.load_pdf_page(self.current_pdf_page + 1)
            
    def show_go_to_page_dialog(self):
        """显示页面跳转对话框"""
        if not self.pdf_file_path or self.pdf_page_count <= 1:
            return
            
        page, ok = QInputDialog.getInt(
            self, 
            "跳转到页面", 
            f"请输入要跳转的页码 (1-{self.pdf_page_count}):", 
            self.current_pdf_page + 1,  # 当前页码（从1开始）
            1, self.pdf_page_count, 1
        )
        
        if ok:
            self.load_pdf_page(page - 1)  # 转换为从0开始的索引

    def setup_compact_ocr_panel(self, parent_layout):
        ocr_widget = QWidget(); ocr_widget.setMaximumHeight(200); ocr_layout = QVBoxLayout(ocr_widget); ocr_layout.setContentsMargins(5, 5, 5, 5); ocr_layout.setSpacing(3)
        row1_layout = QHBoxLayout(); row1_layout.addWidget(QLabel("语言:")); self.language_combo = QComboBox(); self.language_combo.addItems(list(DEFAULT_OCR_LANGUAGES.keys())); self.language_combo.setCurrentText("中文+英文"); row1_layout.addWidget(self.language_combo)
        row1_layout.addWidget(QLabel("置信度:")); self.confidence_slider = QSlider(Qt.Horizontal); self.confidence_slider.setRange(10, 90); self.confidence_slider.setValue(30); self.confidence_slider.setMaximumWidth(80); self.confidence_label = QLabel("0.30"); self.confidence_label.setMinimumWidth(40); row1_layout.addWidget(self.confidence_slider); row1_layout.addWidget(self.confidence_label); ocr_layout.addLayout(row1_layout)
        row2_layout = QHBoxLayout()
        self.enhance_contrast_cb = QCheckBox("增强对比度"); self.enhance_contrast_cb.setChecked(True); row2_layout.addWidget(self.enhance_contrast_cb)
        self.denoise_cb = QCheckBox("降噪"); self.denoise_cb.setChecked(True); row2_layout.addWidget(self.denoise_cb)
        self.gpu_checkbox = QCheckBox("GPU"); self.gpu_checkbox.setChecked(HAS_GPU_SUPPORT); self.gpu_checkbox.setEnabled(HAS_GPU_SUPPORT); row2_layout.addWidget(self.gpu_checkbox)
        self.cpu_checkbox = QCheckBox("CPU"); self.cpu_checkbox.setChecked(not HAS_GPU_SUPPORT); row2_layout.addWidget(self.cpu_checkbox)
        row2_layout.addWidget(QLabel("线程数:"))
        self.threads_spinbox = QSpinBox()
        self.threads_spinbox.setMinimum(1)
        self.threads_spinbox.setMaximum(32)
        self.threads_spinbox.setValue(8)  # 默认8线程
        self.threads_spinbox.setToolTip("CPU模式下使用的线程数")
        self.threads_spinbox.setEnabled(not HAS_GPU_SUPPORT)  # 初始状态根据CPU选择框状态
        row2_layout.addWidget(self.threads_spinbox)
        row2_layout.addStretch(); ocr_layout.addLayout(row2_layout)
        row3_layout = QHBoxLayout(); self.ocr_button = QPushButton("🔍 开始OCR识别" if HAS_OCR_SUPPORT else "❌ OCR不可用");
        if not HAS_OCR_SUPPORT: self.ocr_button.setEnabled(False); self.ocr_button.setToolTip("请安装完整依赖包以启用OCR功能")
        self.ocr_button.setStyleSheet(f"""QPushButton {{ background-color: {UI_COLORS["primary"]}; color: white; font-weight: bold; border: none; min-height: 25px; }} QPushButton:hover {{ background-color: {UI_COLORS["secondary"]}; }} QPushButton:disabled {{ background-color: #cccccc; color: #666666; }}""")
        row3_layout.addWidget(self.ocr_button); self.create_all_btn = QPushButton("全部标注"); self.create_all_btn.setMaximumWidth(80); row3_layout.addWidget(self.create_all_btn); self.clear_ocr_btn = QPushButton("清除OCR"); self.clear_ocr_btn.setMaximumWidth(80); row3_layout.addWidget(self.clear_ocr_btn); ocr_layout.addLayout(row3_layout)
        self.progress_bar = QProgressBar(); self.progress_bar.setVisible(False); self.progress_bar.setMaximumHeight(15); ocr_layout.addWidget(self.progress_bar); self.ocr_stats_label = QLabel("识别结果: 0个文本"); self.ocr_stats_label.setStyleSheet("QLabel { background-color: transparent; border: none; padding: 4px; color: #6c757d; font-size: 11px; }"); ocr_layout.addWidget(self.ocr_stats_label)
        filter_layout = QHBoxLayout(); filter_layout.addWidget(QLabel("筛选:")); self.filter_combo = QComboBox(); self.filter_combo.addItems(OCR_FILTER_OPTIONS); filter_layout.addWidget(self.filter_combo); filter_layout.addStretch(); ocr_layout.addLayout(filter_layout)
        parent_layout.addWidget(ocr_widget)

    def center_view(self):
        """确保图像在视图中居中显示"""
        if not self.graphics_scene.items():
            return
            
        logger.debug("开始执行center_view()方法")
        
        # 直接调用一次centerContent，不使用任何定时器或延迟
        self.graphics_view.centerContent()
        
        # 强制处理所有待处理事件，确保界面更新
        QApplication.processEvents()
        
        logger.debug("center_view()方法执行完成")

    def event(self, event):
        """处理自定义事件"""
        if event.type() == LoadPDFEvent.EVENT_TYPE:
            # 处理PDF加载事件
            try:
                logger.debug(f"接收到PDF加载事件")
                pixmap = event.pixmap
                temp_path = event.temp_path
                
                # 清除并设置场景
                self.graphics_scene.clear()
                self.graphics_scene.addPixmap(pixmap)
                self.current_pixmap = pixmap
                self.current_file_path = temp_path
                
                # 强制处理事件，确保场景已更新
                QApplication.processEvents()
                
                # 仅在初始加载时执行一次居中操作，不重复居中或使用定时器
                self.graphics_view.centerContent()
                
                # 恢复此页面的数据
                self.restore_page_data(self.current_pdf_page)
                
                self.status_bar.showMessage(f"✅ 页面加载成功: {self.current_pdf_page+1}/{self.pdf_page_count}", 3000)
                
                # 隐藏加载对话框
                self.loading_dialog.hide()
                
                logger.debug(f"PDF加载事件处理完成")
                return True
            except Exception as e:
                logger.exception(f"处理PDF加载事件时出错: {str(e)}")
        
        return super().event(event)
    
    def resizeEvent(self, event):
        """窗口大小调整时重新居中图像"""
        super().resizeEvent(event)
        
        # 调整loading对话框大小
        if hasattr(self, 'loading_dialog'):
            self.loading_dialog.resize(self.size())
            
        # 如果没有图像，则不需要居中
        if not self.graphics_scene.items():
            return
        
        # 不再在每次调整窗口大小时都重置缩放，因为这会干扰用户的缩放操作
        logger.debug("窗口大小已改变，但不自动重置缩放以避免干扰用户操作")

    def sync_size_input_from_slider(self):
        """同步滑块值到输入框"""
        self.size_input.setText(str(self.size_slider.value()))

    def on_size_input_changed(self):
        """输入框编辑完成时更新气泡大小"""
        try:
            new_size = int(self.size_input.text())
            self.size_slider.setValue(new_size)
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字。")